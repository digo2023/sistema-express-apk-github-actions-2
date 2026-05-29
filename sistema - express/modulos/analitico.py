import re
import csv
import shutil
import unicodedata
from pathlib import Path
from collections import defaultdict
from datetime import datetime
from xml.sax.saxutils import escape

from configuracao import IMPORT_DIR, PDF_DIR, EXCEL_DIR, ASSETS_DIR, agora_stamp, hoje_br, assinatura_relatorio, AUTHOR, APP_NAME
from modulos.interface import header, ask, warn, ok, tabela, pause, menu, confirm, info, executar_seguro
from modulos.banco_dados import execute, executemany, fetchall, fetchone, evento
from modulos.utilitarios import br_to_iso, iso_to_br, fmt_qtd, pedir_data, pedir_periodo, nome_arquivo_seguro

# ==========================================================
# PADRÃO TEKINISA — REQUISIÇÃO ANALÍTICA POR SERVIÇO
# ==========================================================
# Este módulo lê o PDF padrão PLA00700.QR2 (Requisição Analítica por Serviço),
# filtra por período e turno, soma todos os itens e gera relatórios operacionais.
# Mantém compatibilidade com as tabelas das edições anteriores.


def _limpar(txt):
    return str(txt or '').strip()


def _sem_acento(txt):
    txt = unicodedata.normalize('NFKD', str(txt or ''))
    return ''.join(c for c in txt if not unicodedata.combining(c))


def _chave_nome(txt):
    txt = _sem_acento(txt).upper()
    txt = re.sub(r'[^A-Z0-9]+', ' ', txt)
    txt = re.sub(r'\b(KG|LT|UN|PC|PCT|CX|CAIXA|PACOTE|X|COM|DE|DA|DO|DAS|DOS)\b', ' ', txt)
    return re.sub(r'\s+', ' ', txt).strip()


def normalizar_turno(servico):
    s = _sem_acento(servico).upper()
    if 'ALMOCO' in s:
        return 'A'
    if 'JANTAR' in s or 'JANTA' in s:
        return 'B'
    if 'CEIA' in s:
        return 'C'
    return ''


def nome_turno(t):
    return {'A': 'A - Almoço', 'B': 'B - Janta', 'C': 'C - Ceia'}.get(t, 'Todos')


def nome_dia_semana_iso(data_iso):
    try:
        d = datetime.strptime(str(data_iso)[:10], '%Y-%m-%d')
        return ['segunda-feira', 'terça-feira', 'quarta-feira', 'quinta-feira', 'sexta-feira', 'sábado', 'domingo'][d.weekday()]
    except Exception:
        return ''


def eh_proteina(codigo, item):
    nome = _chave_nome(item)
    cod = str(codigo or '')
    if cod.startswith('1.01.') or cod.startswith('1.11.'):
        return True
    if cod.startswith('1.02.03.') and ('QUEIJO' in nome or 'MUSSARELA' in nome or 'MUCAR' in nome):
        return True
    termos = [
        'CARNE', 'FRANGO', 'BOVINA', 'BOVINO', 'SUINA', 'SUINO', 'LINGUICA', 'CALABRESA',
        'HAMBURGUER', 'SASSAMI', 'OVO', 'QUEIJO', 'MUSSARELA', 'PEIXE', 'TILAPIA',
        'PERNIL', 'COSTELA', 'ACEM', 'PALETA', 'COXA', 'SOBRECOXA', 'SALSICHA', 'LOMBO',
        'ALMONDEGA', 'KIBE', 'NUGGET', 'FILE', 'BIFE', 'MORTADELA', 'PRESUNTO', 'BACON'
    ]
    return any(t in nome for t in termos)


def categoria_item(codigo, item):
    cod = str(codigo or '')
    nome = _chave_nome(item)
    if eh_proteina(codigo, item):
        return 'PROTEINAS'
    if cod.startswith('1.17.') or any(t in nome for t in ['ALFACE', 'TOMATE', 'CEBOLA', 'BATATA', 'CENOURA', 'REPOLHO', 'MELANCIA', 'BANANA', 'HORTIFRUTI', 'ABOBRINHA', 'MELAO']):
        return 'HORTIFRUTI'
    if cod.startswith('3.01.') or any(t in nome for t in ['COPO', 'PRATO DESCARTAVEL', 'GUARDANAPO', 'SACO LIXO', 'BOBINA', 'EMBALAGEM', 'MARMITEX']):
        return 'DESCARTAVEIS'
    if cod.startswith('3.04.') or 'GAS GLP' in nome:
        return 'OPERACIONAIS'
    return 'NAO PERECIVEIS'


def _to_float(valor):
    v = str(valor or '').strip().replace(' ', '')
    if ',' in v:
        v = v.replace('.', '').replace(',', '.')
    try:
        return float(v)
    except Exception:
        return 0.0


def _norm_unidade(un):
    u = _sem_acento(un).upper().strip().replace('.', '')
    mapa = {
        'PC': 'PCT', 'PCT': 'PCT', 'PACOTE': 'PCT',
        'CX': 'CX', 'CAIXA': 'CX',
        'UNID': 'UN', 'UNIDADE': 'UN', 'UN': 'UN',
        'L': 'LT', 'LT': 'LT', 'LITRO': 'LT', 'LITROS': 'LT',
        'KG': 'KG', 'G': 'G', 'ML': 'ML'
    }
    return mapa.get(u, u or 'UN')


def _extrair_texto_pdf(caminho):
    # Usa pdftotext quando disponível: é rápido e preserva melhor as colunas do Tekinisa.
    try:
        import shutil as _shutil, subprocess
        if _shutil.which('pdftotext'):
            res = subprocess.run(['pdftotext', '-layout', str(caminho), '-'], capture_output=True, text=True, timeout=180)
            if res.returncode == 0 and res.stdout.strip():
                return res.stdout
    except Exception:
        pass
    # Fallback puro Python. Pode ser mais lento em PDFs grandes.
    try:
        from pypdf import PdfReader
        reader = PdfReader(caminho)
        return '\n'.join((page.extract_text() or '') for page in reader.pages)
    except Exception as exc:
        raise RuntimeError(
            'Não foi possível ler o PDF. No Termux, instale o leitor leve com: '
            'pkg install poppler -y && pip install -r bibliotecas-analitico.txt'
        ) from exc


def extrair_pdf(caminho):
    """Extrai o PDF padrão Tekinisa Requisição Analítica (por Serviço).

    Campos essenciais salvos: data, serviço, turno, prato, item, quantidade, per capita e unidade.
    O parser lê os itens de trás para frente para não se perder em nomes com números, como
    COPO PLASTICO DESCARTAVEL 300 ML ou CARNE BOVINA HAMBURGUER 90 G.
    """
    texto = _extrair_texto_pdf(caminho)
    linhas_pdf = [l.strip() for l in texto.splitlines() if l.strip()]

    reg_emissao = re.compile(r'Emiss[aã]o\s*:\s*(\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2}:\d{2})', re.I)
    reg_versao = re.compile(r'Vers[aã]o\s*:\s*([\d\.]+\s*WEB)', re.I)
    reg_filial = re.compile(r'Filial\s+(\d+)\s+(.+)$', re.I)
    reg_data = re.compile(r'(?:^|\s)(\d{2}/\d{2}/\d{4})\s+(segunda-feira|terca-feira|terça-feira|quarta-feira|quinta-feira|sexta-feira|sabado|sábado|domingo)', re.I)
    reg_servico = re.compile(r'(?:^|\s)(\d{3,6})\s+((?:ALMOCO|ALMOÇO|JANTAR|JANTA|CEIA).*)$', re.I)
    reg_limite = re.compile(r'Lim\s*ite\s+de\s+Entrega\s+(\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2})', re.I)
    reg_prato = re.compile(r'Prato\s+([\d\.]+)\s+(.+?)(?:\s+Com\s*ensais\s+(\d+[\d\.,]*)\s+Consum\s*o\s+([\d\.,]+)\s*%)?$', re.I)
    reg_item = re.compile(r'^(\d+\.\d+\.\d+\.\d+\.\d+)\s+(.+?)\s+([\d\.,]+)\s+([\d\.,]+)\s+(KG|LT|UN|L|ML|PC|PCT|G|CX)\s*$', re.I)
    reg_item_pypdf = re.compile(r'^(\d+\.\d+\.\d+\.\d+\.\d+)\s+(.+?)\s+([\d\.,]+)\s+(KG|LT|UN|L|ML|PC|PCT|G|CX)\s*([\d\.,]+)\s*$', re.I)

    registros = []
    emissao = ''
    versao = ''
    filial_codigo = ''
    filial_nome = ''
    data_atual = ''
    dia_semana = ''
    codigo_servico = ''
    servico_atual = ''
    turno_atual = ''
    refeicoes_estimadas = 0
    limite_entrega = ''
    codigo_prato = ''
    nome_prato = ''
    comensais_prato = 0
    consumo_percentual = 0.0
    aguardando_refeicoes = False

    for linha in linhas_pdf:
        me = reg_emissao.search(linha)
        if me:
            emissao = me.group(1)
        mv = reg_versao.search(linha)
        if mv:
            versao = mv.group(1).strip()
        mf = reg_filial.search(linha)
        if mf and not filial_codigo:
            filial_codigo = mf.group(1).strip()
            filial_nome = mf.group(2).strip().upper()

        md = reg_data.search(_sem_acento(linha).lower())
        if md:
            data_atual = br_to_iso(md.group(1))
            dia_semana = md.group(2)
            nums = re.findall(r'\b\d{1,5}\b', linha)
            if nums:
                try:
                    refeicoes_estimadas = int(nums[-1])
                except Exception:
                    pass
            continue

        ms = reg_servico.search(linha)
        if ms:
            codigo_servico = ms.group(1).strip()
            servico_atual = _sem_acento(ms.group(2)).upper().strip()
            servico_atual = re.sub(r'\s+RE\s*FE.*$', '', servico_atual, flags=re.I).strip()
            turno_atual = normalizar_turno(servico_atual)
            refeicoes_estimadas = 0
            aguardando_refeicoes = True
            continue

        if aguardando_refeicoes and re.fullmatch(r'\d+', linha):
            try:
                refeicoes_estimadas = int(linha)
            except Exception:
                refeicoes_estimadas = 0
            aguardando_refeicoes = False
            continue

        ml = reg_limite.search(linha)
        if ml:
            limite_entrega = ml.group(1)
            continue

        mp = reg_prato.search(linha)
        if mp:
            codigo_prato = mp.group(1).strip()
            nome_raw = _sem_acento(mp.group(2)).upper().strip()
            mc = re.search(r'COM\s*E?\s*NS\s*AIS\s+(\d+[\d\.,]*)\s+CONS\s*UM\s*O\s+([\d\.,]+)\s*%', nome_raw, re.I)
            if mc:
                comensais_prato = int(_to_float(mc.group(1)))
                consumo_percentual = _to_float(mc.group(2))
                nome_raw = re.sub(r'\s+COM\s*E?\s*NS\s*AIS.*$', '', nome_raw, flags=re.I).strip()
            else:
                comensais_prato = int(_to_float(mp.group(3))) if mp.group(3) else 0
                consumo_percentual = _to_float(mp.group(4)) if mp.group(4) else 0.0
            nome_prato = nome_raw
            continue

        mi = reg_item.match(linha)
        formato_pypdf = False
        if not mi:
            mi = reg_item_pypdf.match(linha)
            formato_pypdf = bool(mi)
        if mi and data_atual:
            if formato_pypdf:
                codigo, item, qtd, unidade, per_capita = mi.groups()
            else:
                codigo, item, qtd, per_capita, unidade = mi.groups()
            item = _sem_acento(_limpar(item)).upper()
            unidade = _norm_unidade(unidade)
            registros.append({
                'arquivo': Path(caminho).name,
                'emissao': emissao,
                'versao': versao,
                'filial_codigo': filial_codigo,
                'filial_nome': filial_nome,
                'data': data_atual,
                'dia_semana': dia_semana,
                'codigo_servico': codigo_servico,
                'servico': servico_atual,
                'turno': turno_atual,
                'refeicoes_estimadas': refeicoes_estimadas,
                'limite_entrega': limite_entrega,
                'codigo_prato': codigo_prato,
                'nome_prato': nome_prato,
                'comensais_prato': comensais_prato,
                'consumo_percentual': consumo_percentual,
                'codigo': codigo,
                'item': item,
                'quantidade': _to_float(qtd),
                'per_capita': _to_float(per_capita),
                'unidade': unidade,
                'categoria': categoria_item(codigo, item),
            })
    return registros


def _datas_turnos_resumo(registros):
    datas = sorted(set(r['data'] for r in registros if r.get('data')))
    turnos = sorted(set(r['turno'] for r in registros if r.get('turno')))
    cats = defaultdict(int)
    for r in registros:
        cats[r['categoria']] += 1
    return datas, turnos, cats


def importar_analitico():
    header('IMPORTAR REQUISIÇÃO ANALÍTICA')
    info('Informe o caminho completo do PDF. Exemplo: /storage/emulated/0/Download/analitico.pdf')
    from modulos.arquivos import selecionar_arquivo
    caminho = selecionar_arquivo('analitico', 'IMPORTAR REQUISIÇÃO ANALÍTICA')
    caminho_path = Path(caminho)
    if not caminho_path.exists() or not caminho_path.is_file():
        warn('Arquivo não encontrado. Confira o caminho e tente novamente.')
        pause(); return
    if caminho_path.suffix.lower() != '.pdf':
        warn('O arquivo informado não parece ser PDF. Confira o caminho antes de continuar.')
        pause(); return
    nome_original = nome_arquivo_seguro(caminho_path.name)
    ja_importado = fetchone('SELECT COUNT(*) AS c FROM analitico_itens WHERE arquivo LIKE ?', (f'%{nome_original}',))
    if ja_importado and int(ja_importado['c'] or 0) > 0:
        warn(f'Este arquivo aparenta já ter sido importado ({ja_importado["c"]} registros encontrados).')
        if not confirm('Importar novamente mesmo assim? Isso pode duplicar quantidades nos relatórios.', default=False):
            warn('Importação cancelada para evitar duplicidade.')
            pause(); return
    try:
        registros = extrair_pdf(caminho_path)
    except Exception as e:
        warn(str(e)); pause(); return
    if not registros:
        warn('Nenhum item foi encontrado. Verifique se o arquivo é a Requisição Analítica Tekinisa correta.')
        pause(); return
    datas, turnos, cats = _datas_turnos_resumo(registros)
    dados_cat = [(c, qtd) for c, qtd in sorted(cats.items())]
    tabela('PRÉVIA DA IMPORTAÇÃO', ['Categoria', 'Registros'], dados_cat)
    print(f"\nArquivo: {Path(caminho).name}")
    print(f"Período encontrado: {iso_to_br(datas[0])} a {iso_to_br(datas[-1])}" if datas else 'Período não identificado')
    print(f"Turnos encontrados: {', '.join(turnos) if turnos else 'não identificado'}")
    print(f"Total de registros encontrados: {len(registros)}")
    if not confirm('Confirmar importação destes dados?', default=True):
        warn('Importação cancelada. Nenhum dado foi salvo.')
        pause(); return
    dest = IMPORT_DIR / f'requisicao_analitica_{agora_stamp()}_{nome_original}'
    try:
        shutil.copy2(caminho, dest)
        nome_arquivo = dest.name
    except Exception:
        nome_arquivo = nome_original
    sql_import = '''INSERT INTO analitico_itens(
                    arquivo,data,servico,turno,codigo,item,quantidade,unidade,categoria,
                    emissao,versao,filial_codigo,filial_nome,dia_semana,codigo_servico,
                    refeicoes_estimadas,limite_entrega,codigo_prato,nome_prato,comensais_prato,
                    consumo_percentual,per_capita)
                   VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)'''
    dados_import = [
        (nome_arquivo, r['data'], r['servico'], r['turno'], r['codigo'], r['item'], r['quantidade'], r['unidade'], r['categoria'],
         r['emissao'], r['versao'], r['filial_codigo'], r['filial_nome'], r['dia_semana'], r['codigo_servico'],
         r['refeicoes_estimadas'], r['limite_entrega'], r['codigo_prato'], r['nome_prato'], r['comensais_prato'],
         r['consumo_percentual'], r['per_capita'])
        for r in registros
    ]
    executemany(sql_import, dados_import)
    evento('PLANEJAMENTO', f'Importada requisição analítica {nome_arquivo} com {len(registros)} registros')
    ok(f'Importação concluída com sucesso: {len(registros)} registros salvos.')
    pause()


def escolher_turnos():
    texto = ask('Turnos [A, B, C, AB, AC, BC ou TODOS]', default='TODOS', upper=True)
    texto = texto.replace(',', '').replace(' ', '')
    if texto in ['', 'TODOS', 'T']:
        return ['A', 'B', 'C']
    turnos = []
    for ch in texto:
        if ch in ['A', 'B', 'C'] and ch not in turnos:
            turnos.append(ch)
    if not turnos:
        warn('Turno inválido. Será considerado TODOS.')
        return ['A', 'B', 'C']
    return turnos


def _filtro_periodo(ini, fim, turnos):
    params = [ini, fim]
    filtro = 'data BETWEEN ? AND ?'
    if turnos and set(turnos) != {'A', 'B', 'C'}:
        ph = ','.join(['?'] * len(turnos))
        filtro += f' AND turno IN ({ph})'
        params.extend(turnos)
    return filtro, params


def _buscar_consolidado(ini, fim, turnos):
    filtro, params = _filtro_periodo(ini, fim, turnos)
    return fetchall(f'''SELECT codigo,categoria,item,unidade,SUM(quantidade) AS qtd, COUNT(*) AS registros,
                              MIN(data) AS primeira_data, MAX(data) AS ultima_data
                        FROM analitico_itens WHERE {filtro}
                        GROUP BY codigo,categoria,item,unidade ORDER BY categoria,item''', params)


def _buscar_detalhe_proteinas(ini, fim, turnos):
    filtro, params = _filtro_periodo(ini, fim, turnos)
    return fetchall(f'''SELECT item,unidade,data,SUM(quantidade) AS qtd
                        FROM analitico_itens WHERE {filtro} AND categoria='PROTEINAS'
                        GROUP BY item,unidade,data ORDER BY item,data''', params)


def _buscar_detalhe_completo(ini, fim, turnos):
    filtro, params = _filtro_periodo(ini, fim, turnos)
    return fetchall(f'''SELECT categoria,item,unidade,data,turno,servico,nome_prato,SUM(quantidade) AS qtd
                        FROM analitico_itens WHERE {filtro}
                        GROUP BY categoria,item,unidade,data,turno,servico,nome_prato
                        ORDER BY data,turno,categoria,item,nome_prato''', params)


def _buscar_uso_por_item(ini, fim, turnos):
    filtro, params = _filtro_periodo(ini, fim, turnos)
    rows = fetchall(f'''SELECT codigo,item,unidade,categoria,data,turno,SUM(quantidade) AS qtd
                        FROM analitico_itens WHERE {filtro}
                        GROUP BY codigo,item,unidade,categoria,data,turno
                        ORDER BY item,data,turno''', params)
    mapa = defaultdict(list)
    for r in rows:
        mapa[(r['codigo'], r['item'], r['unidade'], r['categoria'])].append(r)
    return mapa


def _turnos_txt(turnos):
    if not turnos or set(turnos) == {'A', 'B', 'C'}:
        return 'Todos os turnos: A Almoço, B Janta e C Ceia'
    return ', '.join(nome_turno(t) for t in turnos)


def resumo_analitico():
    header('RESUMO DA REQUISIÇÃO')
    ini, fim = pedir_periodo('Data inicial', 'Data final')
    turnos = escolher_turnos()
    rows = _buscar_consolidado(ini, fim, turnos)
    dados = [(r['categoria'], r['item'], fmt_qtd(r['qtd'], r['unidade']), r['unidade'], iso_to_br(r['primeira_data']), iso_to_br(r['ultima_data'])) for r in rows]
    tabela(f'CONSOLIDADO {iso_to_br(ini)} A {iso_to_br(fim)}', ['Categoria', 'Item', 'Qtd', 'Un', '1º uso', 'Últ. uso'], dados, 300)
    if not rows:
        warn('Nenhum registro encontrado para este período/turno.')
    pause()


def gerar_excel_planejamento(ini, fim, turnos, consolidado, detalhe_prot, detalhe_completo):
    arquivo = EXCEL_DIR / f'uso_periodo_tekinisa_{agora_stamp()}.xlsx'
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    azul = '1F4E78'; vermelho = '8B0000'

    ws = wb.active
    ws.title = 'Resumo por Item'
    ws.append([APP_NAME])
    ws.append([f'Relatório de Uso do Período - {iso_to_br(ini)} a {iso_to_br(fim)} - {assinatura_relatorio()}'])
    ws.append([f'Turnos: {_turnos_txt(turnos)}'])
    ws.append([])
    ws.append(['Categoria', 'Código', 'Item', 'Quantidade', 'Unidade', 'Primeiro uso', 'Último uso', 'Registros'])
    for c in ws[5]:
        c.font = Font(bold=True, color='FFFFFF'); c.fill = PatternFill('solid', fgColor=azul); c.alignment = Alignment(horizontal='center')
    for r in consolidado:
        ws.append([r['categoria'], r['codigo'], r['item'], float(r['qtd'] or 0), r['unidade'], iso_to_br(r['primeira_data']), iso_to_br(r['ultima_data']), r['registros']])
    for i, w in enumerate([22, 18, 52, 15, 10, 14, 14, 12], 1):
        ws.column_dimensions[chr(64 + i)].width = w

    ws2 = wb.create_sheet('Proteínas por dia')
    ws2.append(['Proteína', 'Data', 'Dia da semana', 'Quantidade', 'Unidade'])
    for c in ws2[1]:
        c.font = Font(bold=True, color='FFFFFF'); c.fill = PatternFill('solid', fgColor=vermelho); c.alignment = Alignment(horizontal='center')
    for r in detalhe_prot:
        ws2.append([r['item'], iso_to_br(r['data']), nome_dia_semana_iso(r['data']), float(r['qtd'] or 0), r['unidade']])
    for i, w in enumerate([52, 14, 20, 15, 12], 1):
        ws2.column_dimensions[chr(64 + i)].width = w

    ws3 = wb.create_sheet('Uso detalhado')
    ws3.append(['Categoria', 'Item', 'Data', 'Turno', 'Serviço', 'Prato', 'Quantidade', 'Unidade'])
    for c in ws3[1]:
        c.font = Font(bold=True, color='FFFFFF'); c.fill = PatternFill('solid', fgColor=azul); c.alignment = Alignment(horizontal='center')
    for r in detalhe_completo:
        ws3.append([r['categoria'], r['item'], iso_to_br(r['data']), r['turno'], r['servico'], r['nome_prato'], float(r['qtd'] or 0), r['unidade']])
    for i, w in enumerate([22, 52, 14, 10, 36, 52, 15, 12], 1):
        ws3.column_dimensions[chr(64 + i)].width = w

    for wsx in wb.worksheets:
        wsx.freeze_panes = 'A6' if wsx.title == 'Resumo por Item' else 'A2'
    wb.save(arquivo)
    return arquivo


def gerar_pdf_planejamento(ini, fim, turnos, consolidado, detalhe_prot):
    arquivo = PDF_DIR / f'uso_periodo_tekinisa_{agora_stamp()}.pdf'
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, KeepTogether, Image
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.units import cm
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='TituloCentro', parent=styles['Title'], alignment=TA_CENTER, fontSize=15, leading=18))
    styles.add(ParagraphStyle(name='Subtitulo', parent=styles['BodyText'], alignment=TA_CENTER, fontSize=9, leading=12, spaceAfter=8))
    styles.add(ParagraphStyle(name='Secao', parent=styles['Heading2'], fontSize=12, leading=14, spaceBefore=6, spaceAfter=6))
    styles.add(ParagraphStyle(name='Pequeno', parent=styles['BodyText'], fontSize=7.4, leading=9))
    def p(txt, style='Pequeno'):
        return Paragraph(escape(str(txt)), styles[style])
    doc = SimpleDocTemplate(str(arquivo), pagesize=landscape(A4), rightMargin=.7*cm, leftMargin=.7*cm, topMargin=.7*cm, bottomMargin=1.1*cm)
    story = []
    logo = ASSETS_DIR / 'logo.png'
    if logo.exists():
        try:
            img = Image(str(logo), width=2.6*cm, height=2.6*cm); img.hAlign = 'CENTER'; story.append(img); story.append(Spacer(1, 3))
        except Exception:
            pass
    story.append(p(f'RELATÓRIO DE USO DO PERÍODO - PADRÃO TEKINISA', 'TituloCentro'))
    story.append(p(f'{iso_to_br(ini)} a {iso_to_br(fim)} | {_turnos_txt(turnos)}', 'Subtitulo'))
    total_reg = sum(int(r['registros'] or 0) for r in consolidado)
    prot_unicas = len([r for r in consolidado if r['categoria'] == 'PROTEINAS'])
    info_rows = [['Período', f'{iso_to_br(ini)} a {iso_to_br(fim)}'], ['Turnos', _turnos_txt(turnos)], ['Registros considerados', str(total_reg)], ['Proteínas únicas', str(prot_unicas)], ['Emissão', assinatura_relatorio()]]
    t_info = Table([[p(a), p(b)] for a, b in info_rows], colWidths=[5.2*cm, 20.8*cm])
    t_info.setStyle(TableStyle([('GRID', (0,0), (-1,-1), .25, colors.grey), ('BACKGROUND', (0,0), (0,-1), colors.HexColor('#D9EAF7')), ('VALIGN', (0,0), (-1,-1), 'TOP')]))
    story.append(t_info); story.append(Spacer(1, 8))

    story.append(p('1. RESUMO TOTAL DAS PROTEÍNAS', 'Secao'))
    prot = [r for r in consolidado if r['categoria'] == 'PROTEINAS']
    data = [[p('CÓDIGO'), p('PROTEÍNA'), p('UN'), p('TOTAL'), p('PRIMEIRO USO'), p('ÚLTIMO USO')]]
    for r in prot:
        data.append([p(r['codigo']), p(r['item']), p(r['unidade']), p(fmt_qtd(r['qtd'], r['unidade'])), p(iso_to_br(r['primeira_data'])), p(iso_to_br(r['ultima_data']))])
    if len(data) == 1:
        data.append([p('-'), p('Nenhuma proteína encontrada'), p('-'), p('-'), p('-'), p('-')])
    t = Table(data, colWidths=[3.0*cm, 13.2*cm, 1.4*cm, 3.0*cm, 2.4*cm, 2.4*cm], repeatRows=1)
    t.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,0), colors.HexColor('#8B0000')), ('TEXTCOLOR', (0,0), (-1,0), colors.white), ('GRID', (0,0), (-1,-1), .25, colors.black), ('ALIGN', (2,1), (-1,-1), 'CENTER'), ('VALIGN', (0,0), (-1,-1), 'TOP')]))
    story.append(t); story.append(PageBreak())

    story.append(p('2. DIAS DE USO DAS PROTEÍNAS', 'Secao'))
    agrup = defaultdict(list)
    for r in detalhe_prot:
        agrup[(r['item'], r['unidade'])].append(r)
    if not agrup:
        story.append(p('Nenhuma proteína encontrada no período informado.'))
    for (item, unidade), regs in agrup.items():
        total = sum(float(x['qtd'] or 0) for x in regs)
        bloco = [p(f'<b>{item}</b> - Total: <b>{fmt_qtd(total, unidade)} {unidade}</b>', 'Pequeno')]
        tab = [[p('DATA'), p('DIA'), p('QTDE'), p('UN')]]
        for r in regs:
            tab.append([p(iso_to_br(r['data'])), p(nome_dia_semana_iso(r['data'])), p(fmt_qtd(r['qtd'], unidade)), p(unidade)])
        tb = Table(tab, colWidths=[3.0*cm, 5.0*cm, 3*cm, 1.5*cm], repeatRows=1)
        tb.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,0), colors.HexColor('#F4CCCC')), ('GRID', (0,0), (-1,-1), .25, colors.grey), ('ALIGN', (2,1), (-1,-1), 'CENTER'), ('VALIGN', (0,0), (-1,-1), 'TOP')]))
        bloco.append(tb); bloco.append(Spacer(1, 5)); story.append(KeepTogether(bloco))
    story.append(PageBreak())

    story.append(p('3. LISTA COMPLETA CONSOLIDADA POR CATEGORIA', 'Secao'))
    for cat in ['PROTEINAS', 'NAO PERECIVEIS', 'HORTIFRUTI', 'DESCARTAVEIS', 'OPERACIONAIS']:
        sub = [r for r in consolidado if r['categoria'] == cat]
        if not sub:
            continue
        story.append(p(f'{cat} - Total de itens: {len(sub)}', 'Secao'))
        tab = [[p('CÓDIGO'), p('ITEM'), p('UN'), p('TOTAL'), p('1º USO'), p('ÚLT. USO')]]
        for r in sub:
            tab.append([p(r['codigo']), p(r['item']), p(r['unidade']), p(fmt_qtd(r['qtd'], r['unidade'])), p(iso_to_br(r['primeira_data'])), p(iso_to_br(r['ultima_data']))])
        tb = Table(tab, colWidths=[3*cm, 14.2*cm, 1.3*cm, 2.6*cm, 2.2*cm, 2.2*cm], repeatRows=1)
        tb.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F4E78')), ('TEXTCOLOR', (0,0), (-1,0), colors.white), ('GRID', (0,0), (-1,-1), .25, colors.black), ('ALIGN', (2,1), (-1,-1), 'CENTER'), ('VALIGN', (0,0), (-1,-1), 'TOP')]))
        story.append(tb); story.append(Spacer(1, 8))
    def footer(canvas, doc):
        canvas.saveState(); canvas.setFont('Helvetica', 7); canvas.drawRightString(28.5*cm, .45*cm, f'{assinatura_relatorio()} | Página {doc.page}'); canvas.restoreState()
    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    return arquivo


def gerar_relatorio_planejamento():
    header('GERAR USO DO PERÍODO')
    ini, fim = pedir_periodo('Data inicial', 'Data final')
    turnos = escolher_turnos()
    consolidado = _buscar_consolidado(ini, fim, turnos)
    if not consolidado:
        warn('Nenhum dado importado para o período informado. Primeiro importe a Requisição Analítica.')
        pause(); return
    detalhe_prot = _buscar_detalhe_proteinas(ini, fim, turnos)
    detalhe_completo = _buscar_detalhe_completo(ini, fim, turnos)
    dados = [(r['categoria'], r['item'], fmt_qtd(r['qtd'], r['unidade']), r['unidade'], iso_to_br(r['primeira_data'])) for r in consolidado]
    tabela('PRÉVIA DO RELATÓRIO', ['Categoria', 'Item', 'Qtd', 'Un', '1º uso'], dados, 25)
    if not confirm('Gerar PDF, Excel e CSV deste período?', default=True):
        warn('Geração cancelada.'); pause(); return
    try:
        pdf = gerar_pdf_planejamento(ini, fim, turnos, consolidado, detalhe_prot)
        excel = gerar_excel_planejamento(ini, fim, turnos, consolidado, detalhe_prot, detalhe_completo)
        csv_path = EXCEL_DIR / f'uso_periodo_tekinisa_{agora_stamp()}.csv'
        with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
            w = csv.writer(f, delimiter=';')
            w.writerow(['Relatorio', 'Uso do Periodo Tekinisa'])
            w.writerow(['Emissao', assinatura_relatorio()])
            w.writerow([])
            w.writerow(['Categoria', 'Codigo', 'Item', 'Quantidade', 'Unidade', 'Primeiro uso', 'Ultimo uso', 'Registros'])
            for r in consolidado:
                w.writerow([r['categoria'], r['codigo'], r['item'], fmt_qtd(r['qtd'], r['unidade']), r['unidade'], iso_to_br(r['primeira_data']), iso_to_br(r['ultima_data']), r['registros']])
        ok(f'PDF gerado: {pdf}')
        ok(f'Excel gerado: {excel}')
        ok(f'CSV gerado: {csv_path}')
        evento('RELATORIO', f'Uso do período gerado para {iso_to_br(ini)} a {iso_to_br(fim)}')
    except Exception as e:
        warn(f'Falha ao gerar relatório: {e}')
    pause()


def _produtos_estoque():
    return fetchall('''SELECT p.id,p.codigo,p.nome,p.unidade,p.categoria,COALESCE(SUM(e.quantidade),0) AS total
                       FROM produtos p LEFT JOIN estoque_locais e ON e.produto_id=p.id
                       WHERE p.ativo=1 GROUP BY p.id ORDER BY p.nome''')


def _encontrar_produto_para_item(item, codigo=''):
    item_key = _chave_nome(item)
    rows = _produtos_estoque()
    for p in rows:
        if codigo and str(p['codigo']).strip() == str(codigo).strip():
            return p
    for p in rows:
        if _chave_nome(p['nome']) == item_key:
            return p
    candidatos = []
    for p in rows:
        pk = _chave_nome(p['nome'])
        if item_key and (item_key in pk or pk in item_key):
            candidatos.append(p)
    if candidatos:
        return sorted(candidatos, key=lambda x: len(_chave_nome(x['nome'])))[0]
    return None


def _resultado_necessario_estoque(ini, fim, turnos):
    consolidado = _buscar_consolidado(ini, fim, turnos)
    usos = _buscar_uso_por_item(ini, fim, turnos)
    resultado = []
    for r in consolidado:
        prod = _encontrar_produto_para_item(r['item'], r['codigo'])
        necessario = float(r['qtd'] or 0)
        estoque = float(prod['total'] or 0) if prod else 0.0
        saldo = estoque - necessario
        chave = (r['codigo'], r['item'], r['unidade'], r['categoria'])
        datas_uso = sorted(set(x['data'] for x in usos.get(chave, []) if x['data']))
        if not prod:
            status, criticidade = 'NÃO CADASTRADO', 'ALTA'
        elif saldo < 0:
            status = 'FALTANTE'
            criticidade = 'ALTA' if abs(saldo) >= max(necessario * 0.25, 1) else 'MÉDIA'
        elif saldo <= max(necessario * 0.10, 1):
            status, criticidade = 'NO LIMITE', 'MÉDIA'
        else:
            status, criticidade = 'OK', 'BAIXA'
        resultado.append({'categoria': r['categoria'], 'codigo_analitico': r['codigo'], 'item': r['item'], 'unidade': r['unidade'], 'necessario': necessario, 'estoque': estoque, 'saldo': saldo, 'status': status, 'criticidade': criticidade, 'produto': prod['nome'] if prod else '', 'codigo': prod['codigo'] if prod else '', 'primeira_data': r['primeira_data'], 'ultima_data': r['ultima_data'], 'datas_uso': datas_uso})
    return resultado


def comparar_necessario_estoque():
    header('CONFERÊNCIA NECESSÁRIO X ESTOQUE')
    ini, fim = pedir_periodo('Data inicial', 'Data final')
    turnos = escolher_turnos()
    resultado = _resultado_necessario_estoque(ini, fim, turnos)
    if not resultado:
        warn('Nenhum dado encontrado. Importe o analítico antes de comparar.')
        pause(); return
    dados = [(x['status'], x['criticidade'], x['categoria'], x['item'], fmt_qtd(x['necessario'], x['unidade']), fmt_qtd(x['estoque'], x['unidade']), fmt_qtd(x['saldo'], x['unidade']), x['unidade']) for x in resultado]
    tabela('RESULTADO DA CONFERÊNCIA', ['Status', 'Critic.', 'Categoria', 'Item', 'Necessário', 'Estoque', 'Saldo', 'Un'], dados, 300)
    if confirm('Gerar Excel da conferência necessário x estoque?', default=True):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            arq = EXCEL_DIR / f'conferencia_necessario_estoque_{agora_stamp()}.xlsx'
            wb = Workbook(); ws = wb.active; ws.title = 'Conferência'
            ws.append([APP_NAME]); ws.append([f'Necessário x Estoque - {iso_to_br(ini)} a {iso_to_br(fim)} - {assinatura_relatorio()}']); ws.append([f'Turnos: {_turnos_txt(turnos)}']); ws.append([])
            headers = ['Status', 'Criticidade', 'Categoria', 'Item Analítico', 'Produto no Estoque', 'Código', 'Necessário', 'Estoque', 'Saldo', 'Unidade', 'Datas de consumo']
            ws.append(headers)
            for c in ws[5]:
                c.font = Font(bold=True, color='FFFFFF'); c.fill = PatternFill('solid', fgColor='1F4E78'); c.alignment = Alignment(horizontal='center')
            for x in resultado:
                ws.append([x['status'], x['criticidade'], x['categoria'], x['item'], x['produto'], x['codigo'], x['necessario'], x['estoque'], x['saldo'], x['unidade'], ', '.join(iso_to_br(d) for d in x['datas_uso'])])
            for i, w in enumerate([16, 14, 20, 48, 48, 16, 14, 14, 14, 12, 45], 1):
                ws.column_dimensions[chr(64 + i)].width = w
            wb.save(arq); ok(f'Excel gerado: {arq}')
        except Exception as e:
            warn(f'Erro ao gerar Excel: {e}')
    pause()


def relatorio_faltantes_pdf_excel():
    header('RELATÓRIO DE FALTANTES')
    ini, fim = pedir_periodo('Data inicial', 'Data final')
    turnos = escolher_turnos()
    resultado = _resultado_necessario_estoque(ini, fim, turnos)
    criticos = [x for x in resultado if x['status'] in ('FALTANTE','NÃO CADASTRADO','NO LIMITE')]
    if not resultado:
        warn('Nenhum dado encontrado. Importe o analítico antes de gerar o relatório.')
        pause(); return
    dados = [(x['status'], x['criticidade'], x['categoria'], x['item'], fmt_qtd(x['necessario'], x['unidade']), fmt_qtd(x['estoque'], x['unidade']), fmt_qtd(x['saldo'], x['unidade']), x['unidade']) for x in criticos]
    tabela('ITENS CRÍTICOS / FALTANTES', ['Status','Critic.','Categoria','Item','Necessário','Estoque','Saldo','Un'], dados, 300)
    if not confirm('Gerar PDF e Excel profissional de faltantes?', default=True):
        pause(); return
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        arq_x = EXCEL_DIR / f'relatorio_faltantes_{agora_stamp()}.xlsx'
        wb = Workbook(); ws = wb.active; ws.title = 'Faltantes'
        ws.append([APP_NAME]); ws.append([f'Relatório de Itens Faltantes - {iso_to_br(ini)} a {iso_to_br(fim)} - {assinatura_relatorio()}']); ws.append([f'Turnos: {_turnos_txt(turnos)}']); ws.append([])
        headers = ['Status','Criticidade','Categoria','Item','Produto cadastrado','Código','Necessário','Estoque','Faltante/Saldo','Unidade','Datas de consumo','Observação']
        ws.append(headers)
        for c in ws[5]:
            c.font = Font(bold=True, color='FFFFFF'); c.fill = PatternFill('solid', fgColor='1F4E78'); c.alignment = Alignment(horizontal='center')
        for x in criticos:
            obs = 'Cadastrar produto no estoque' if x['status']=='NÃO CADASTRADO' else ('Comprar/solicitar complemento' if x['saldo'] < 0 else 'Acompanhar consumo')
            ws.append([x['status'],x['criticidade'],x['categoria'],x['item'],x['produto'],x['codigo'],x['necessario'],x['estoque'],x['saldo'],x['unidade'], ', '.join(iso_to_br(d) for d in x['datas_uso']), obs])
        for i,w in enumerate([18,14,20,48,48,16,14,14,14,12,45,35],1):
            ws.column_dimensions[chr(64+i)].width = w
        wb.save(arq_x); ok(f'Excel gerado: {arq_x}')
    except Exception as e:
        warn(f'Erro ao gerar Excel: {e}')
    try:
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
        arq_p = PDF_DIR / f'relatorio_faltantes_{agora_stamp()}.pdf'
        styles = getSampleStyleSheet()
        doc = SimpleDocTemplate(str(arq_p), pagesize=landscape(A4), leftMargin=.7*cm, rightMargin=.7*cm, topMargin=.7*cm, bottomMargin=1*cm)
        story = [Paragraph(APP_NAME, styles['Title']), Paragraph(f'Relatório de Itens Faltantes - {iso_to_br(ini)} a {iso_to_br(fim)}', styles['Normal']), Paragraph(f'Turnos: {_turnos_txt(turnos)} - {assinatura_relatorio()}', styles['Normal']), Spacer(1,8)]
        data = [['Status','Crit.','Categoria','Item','Nec.','Estoque','Saldo','Un','Datas de consumo','Obs']]
        for x in criticos:
            obs = 'Cadastrar' if x['status']=='NÃO CADASTRADO' else ('Complementar' if x['saldo'] < 0 else 'Acompanhar')
            data.append([x['status'], x['criticidade'], x['categoria'], x['item'][:38], fmt_qtd(x['necessario'],x['unidade']), fmt_qtd(x['estoque'],x['unidade']), fmt_qtd(x['saldo'],x['unidade']), x['unidade'], ', '.join(iso_to_br(d) for d in x['datas_uso'])[:38], obs])
        if len(data)==1: data.append(['OK','BAIXA','-','Nenhum item crítico encontrado','-','-','-','-','-','Operação suficiente'])
        tb = Table(data, repeatRows=1, colWidths=[2.8*cm,1.8*cm,3*cm,7.5*cm,2.2*cm,2.2*cm,2.2*cm,1.0*cm,4.0*cm,2.4*cm])
        tb.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1F4E78')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('GRID',(0,0),(-1,-1),.25,colors.grey),('FONTSIZE',(0,0),(-1,-1),6.8),('VALIGN',(0,0),(-1,-1),'MIDDLE')]))
        story.append(tb)
        def footer(canvas, doc):
            canvas.saveState(); canvas.setFont('Helvetica',8); canvas.drawRightString(28.5*cm, .45*cm, assinatura_relatorio()); canvas.restoreState()
        doc.build(story, onFirstPage=footer, onLaterPages=footer); ok(f'PDF gerado: {arq_p}')
    except Exception as e:
        warn(f'Erro ao gerar PDF: {e}')
    evento('RELATORIO', f'Relatório de faltantes gerado para {iso_to_br(ini)} a {iso_to_br(fim)}')
    pause()


def gerar_csv_proteinas():
    header('CSV DE PROTEÍNAS')
    ini, fim = pedir_periodo('Data inicial', 'Data final')
    turnos = escolher_turnos()
    filtro, params = _filtro_periodo(ini, fim, turnos)
    rows = fetchall(f'''SELECT item,unidade,SUM(quantidade) AS qtd FROM analitico_itens
                        WHERE {filtro} AND categoria='PROTEINAS'
                        GROUP BY item,unidade ORDER BY item''', params)
    if not rows:
        warn('Nenhuma proteína encontrada no período.'); pause(); return
    arq = EXCEL_DIR / f'proteinas_{agora_stamp()}.csv'
    with open(arq, 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.writer(f, delimiter=';')
        w.writerow(['Relatorio', 'Proteinas do Periodo'])
        w.writerow(['Emissao', assinatura_relatorio()])
        w.writerow([])
        w.writerow(['Proteína', 'Quantidade', 'Unidade'])
        for r in rows:
            w.writerow([r['item'], fmt_qtd(r['qtd'], r['unidade']), r['unidade']])
    ok(f'CSV gerado: {arq}')
    pause()


def historico_importacoes():
    header('HISTÓRICO DE IMPORTAÇÕES')
    rows = fetchall('''SELECT arquivo, MIN(data) AS ini, MAX(data) AS fim, COUNT(*) AS registros,
                              SUM(CASE WHEN categoria='PROTEINAS' THEN 1 ELSE 0 END) AS prot
                       FROM analitico_itens GROUP BY arquivo ORDER BY MAX(criado_em) DESC''')
    dados = [(r['arquivo'], iso_to_br(r['ini']), iso_to_br(r['fim']), r['registros'], r['prot']) for r in rows]
    tabela('ARQUIVOS IMPORTADOS', ['Arquivo', 'Início', 'Fim', 'Registros', 'Reg. proteínas'], dados, 200)
    pause()


def menu_analitico():
    while True:
        op = menu('PLANEJAMENTO DE PRODUÇÃO', [
            ('1', 'Importar Requisição Analítica Tekinisa'),
            ('2', 'Gerar Relatório de Necessários do Período'),
            ('3', 'Importar Ordem de Compra Tekinisa'),
            ('4', 'Conferir Necessários x Ordem de Compra'),
            ('5', 'Gerar Relatório de Itens Faltantes e Divergentes'),
            ('6', 'Consultar Margem com Estoque Atual'),
            ('7', 'Ver resumo por período e turno'),
            ('8', 'Gerar CSV de proteínas'),
            ('9', 'Histórico de importações'),
            ('0', 'Voltar')
        ])
        if op == '1': importar_analitico()
        elif op == '2': gerar_relatorio_planejamento()
        elif op == '3':
            from modulos.ordem_compra import importar_oc_pdf
            importar_oc_pdf()
        elif op == '4':
            from modulos.ordem_compra import conferir_compra_necessidade
            conferir_compra_necessidade()
        elif op == '5':
            from modulos.ordem_compra import conferir_compra_necessidade
            conferir_compra_necessidade()
        elif op == '6': comparar_necessario_estoque()
        elif op == '7': resumo_analitico()
        elif op == '8': gerar_csv_proteinas()
        elif op == '9': historico_importacoes()
        elif op == '0': break
        else:
            warn('Opção inválida.'); pause()
