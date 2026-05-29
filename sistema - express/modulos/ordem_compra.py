from pathlib import Path
import csv
import re
import shutil
import unicodedata
from collections import defaultdict
from difflib import SequenceMatcher
from datetime import datetime
from xml.sax.saxutils import escape

from modulos.interface import header, ask, ask_float, warn, ok, tabela, pause, menu, confirm, info, executar_seguro
from modulos.banco_dados import execute, executemany, fetchall, fetchone, evento
from modulos.utilitarios import pedir_data, pedir_periodo, iso_to_br, fmt_qtd, br_to_iso, nome_arquivo_seguro
from modulos.analitico import _chave_nome, _resultado_necessario_estoque, escolher_turnos, _turnos_txt
from configuracao import EXCEL_DIR, PDF_DIR, IMPORT_DIR, APP_NAME, AUTHOR, agora_stamp, hoje_br, assinatura_relatorio

# ==========================================================
# PADRÃO TEKINISA — SOLICITAÇÃO DE COMPRA / OC
# ==========================================================
# Lê o PDF CMP05002.QR2: Solicitação de Compra de Produtos Comprados na Matriz.
# Depois cruza a OC com a necessidade gerada pelo Analítico.


def _sem_acento(txt):
    txt = unicodedata.normalize('NFKD', str(txt or ''))
    return ''.join(c for c in txt if not unicodedata.combining(c))


def _numero(valor):
    texto = str(valor or '0').strip().replace(' ', '')
    if ',' in texto:
        texto = texto.replace('.', '').replace(',', '.')
    try:
        return float(texto)
    except Exception:
        return 0.0


def _data(valor):
    texto = str(valor or '').strip()
    return br_to_iso(texto) or texto


def _norm_unidade(un):
    u = _sem_acento(un).upper().strip().replace('.', '')
    mapa = {'PC': 'PCT', 'PCT': 'PCT', 'PACOTE': 'PCT', 'CX': 'CX', 'CAIXA': 'CX', 'UNID': 'UN', 'UNIDADE': 'UN', 'L': 'LT', 'LT': 'LT', 'LITRO': 'LT', 'KG': 'KG', 'G': 'G', 'ML': 'ML'}
    return mapa.get(u, u or 'UN')


def _extrair_texto_pdf(caminho):
    # Usa pdftotext quando disponível: é rápido e preserva melhor as colunas do Tekinisa.
    try:
        import shutil as _shutil, subprocess
        if _shutil.which('pdftotext'):
            res = subprocess.run(['pdftotext', '-layout', str(caminho), '-'], capture_output=True, text=True, timeout=180)
            if res.returncode == 0 and res.stdout.strip():
                return res.stdout
    except Exception:
        pass
    # Fallback puro Python. Pode ser mais lento em PDFs grandes.
    try:
        from pypdf import PdfReader
        reader = PdfReader(caminho)
        return '\n'.join((page.extract_text() or '') for page in reader.pages)
    except Exception as exc:
        raise RuntimeError(
            'Não foi possível ler o PDF. No Termux, instale o leitor leve com: '
            'pkg install poppler -y && pip install -r bibliotecas-analitico.txt'
        ) from exc


def extrair_oc_pdf_tekinisa(caminho):
    """Extrai o PDF padrão Tekinisa de Solicitação de Compra.

    Estrutura esperada por linha de item:
    1.04.01.030.00 ARROZ BRANCO KG 210,000 19/05/2026 21/05/2026
    """
    texto = _extrair_texto_pdf(caminho)
    linhas = [l.strip() for l in texto.splitlines() if l.strip()]

    reg_emissao = re.compile(r'Emiss[aã]o\s*:\s*(\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2}:\d{2})', re.I)
    reg_versao = re.compile(r'Vers[aã]o\s*:\s*([\d\.]+\s*WEB)', re.I)
    reg_filial = re.compile(r'Filial\s+(\d+)\s+(.+)$', re.I)
    reg_periodo = re.compile(r'Per[ií]odo\s+de\s+Entrega\s+(\d{2}/\d{2}/\d{4})\s+a\s+(\d{2}/\d{2}/\d{4})', re.I)
    reg_solic = re.compile(r'Solicita[cç][aã]o\s+(\d+)\s+(\d+).*?(?:Data\s+(\d{2}/\d{2}/\d{4}).*?Descri[cç][aã]o\s+(.+?)|Descri[cç][aã]o\s+(.+?)\s+Destino.*?Data\s+(\d{2}/\d{2}/\d{4}))', re.I)
    reg_item = re.compile(r'^(\d+\.\d+\.\d+\.\d+\.\d+)\s+(.+?)\s+(KG|CX|PC|PCT|UN|LT|L|ML|G)\s+([\d\.,]+)\s+(\d{2}/\d{2}/\d{4})\s+(\d{2}/\d{2}/\d{4})\s*$', re.I)

    emissao = ''
    versao = ''
    filial_codigo = ''
    filial_nome = ''
    periodo_ini = ''
    periodo_fim = ''
    solicitacao = ''
    seq_solic = ''
    data_solic = ''
    descricao = ''
    itens = []

    for linha in linhas:
        me = reg_emissao.search(linha)
        if me:
            emissao = me.group(1)
        mv = reg_versao.search(linha)
        if mv:
            versao = mv.group(1).strip()
        mf = reg_filial.search(linha)
        if mf and not filial_codigo:
            filial_codigo = mf.group(1).strip()
            filial_nome = mf.group(2).strip().upper()
        mp = reg_periodo.search(linha)
        if mp:
            periodo_ini = _data(mp.group(1)); periodo_fim = _data(mp.group(2))
        ms = reg_solic.search(linha)
        if ms:
            solicitacao = ms.group(1).strip()
            seq_solic = ms.group(2).strip()
            data_solic = _data(ms.group(3) or ms.group(6) or '')
            descricao = _sem_acento(ms.group(4) or ms.group(5) or '').upper().strip()
            continue
        if 'Solicita' in linha and re.search(r'\d{8,}', linha):
            nums = re.findall(r'\d+', linha)
            if nums:
                solicitacao = nums[0]
                seq_solic = nums[1] if len(nums) > 1 else ''
            md_s = re.search(r'(\d{2}/\d{2}/\d{4})', linha)
            data_solic = _data(md_s.group(1)) if md_s else ''
            if 'AUTOM' in _sem_acento(linha).upper():
                descricao = 'AUTOMATICA'
            continue
        mi = reg_item.match(linha)
        if mi:
            codigo, item, unidade, qtd, entrega, utilizacao = mi.groups()
            itens.append({
                'documento': f'OC {solicitacao}' if solicitacao else Path(caminho).stem.upper(),
                'codigo': codigo,
                'item': _sem_acento(item).upper().strip(),
                'unidade': _norm_unidade(unidade),
                'quantidade': _numero(qtd),
                'data_entrega': _data(entrega),
                'data_utilizacao': _data(utilizacao),
                'emissao': emissao,
                'versao': versao,
                'filial_codigo': filial_codigo,
                'filial_nome': filial_nome,
                'periodo_entrega_inicio': periodo_ini,
                'periodo_entrega_fim': periodo_fim,
                'solicitacao': solicitacao,
                'sequencia': seq_solic,
                'data_solicitacao': data_solic,
                'descricao': descricao,
            })
    return itens


def cadastrar_item_oc():
    header('ORDEM DE COMPRA / SOLICITAÇÃO')
    documento = ask('Documento/OC', default='OC', upper=True)
    item = ask('Item', upper=True, required=True)
    unidade = ask('Unidade', default='KG', upper=True)
    qtd = ask_float('Quantidade', min_value=0)
    entrega = pedir_data('Data de entrega')
    utilizacao = pedir_data('Data de utilização/consumo')
    obs = ask('Observação', default='')
    print(f"\n{item} | {fmt_qtd(qtd, unidade)} {unidade} | Entrega {iso_to_br(entrega)} | Utilização {iso_to_br(utilizacao)}")
    if not confirm('Confirmar cadastro deste item de OC?', default=True):
        warn('Cancelado.'); pause(); return
    execute('''INSERT INTO oc_itens(documento,data_entrega,data_utilizacao,item,quantidade,unidade,observacao)
               VALUES(?,?,?,?,?,?,?)''', (documento, entrega, utilizacao, item, qtd, unidade, obs))
    evento('OC', f'Item cadastrado na OC: {item} {fmt_qtd(qtd, unidade)} {unidade}')
    ok('Item da OC salvo.')
    pause()


def importar_oc_csv():
    header('IMPORTAR OC EM CSV')
    info('CSV esperado: item;quantidade;unidade;entrega;utilizacao;documento;observacao')
    from modulos.arquivos import selecionar_arquivo
    caminho = selecionar_arquivo('oc', 'IMPORTAR OC EM CSV')
    if not Path(caminho).exists():
        warn('Arquivo não encontrado.'); pause(); return
    dados_import = []
    with open(caminho, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f, delimiter=';')
        for row in reader:
            item = (row.get('item') or row.get('produto') or '').upper().strip()
            if not item:
                continue
            qtd = _numero(row.get('quantidade'))
            unidade = _norm_unidade(row.get('unidade') or 'KG')
            entrega = _data(row.get('entrega') or row.get('data_entrega') or '')
            utilizacao = _data(row.get('utilizacao') or row.get('data_utilizacao') or '')
            documento = (row.get('documento') or 'OC_CSV').upper().strip()
            codigo = (row.get('codigo') or '').strip()
            obs = row.get('observacao') or ''
            dados_import.append((documento, entrega, utilizacao, codigo, item, qtd, unidade, obs))
    total = executemany('''INSERT INTO oc_itens(documento,data_entrega,data_utilizacao,codigo,item,quantidade,unidade,observacao)
                       VALUES(?,?,?,?,?,?,?,?)''', dados_import) if dados_import else 0
    evento('OC', f'Importação CSV de OC: {total} itens')
    ok(f'Importação concluída: {total} itens.')
    pause()


def importar_oc_pdf():
    header('IMPORTAR SOLICITAÇÃO DE COMPRA TEKINISA')
    info('Informe o caminho completo do PDF da Solicitação de Compra. Exemplo: /storage/emulated/0/Download/rancho.pdf')
    from modulos.arquivos import selecionar_arquivo
    caminho = selecionar_arquivo('oc', 'IMPORTAR SOLICITAÇÃO DE COMPRA')
    caminho_path = Path(caminho)
    if not caminho_path.exists() or not caminho_path.is_file():
        warn('Arquivo não encontrado.'); pause(); return
    if caminho_path.suffix.lower() != '.pdf':
        warn('O arquivo informado não parece ser PDF. Confira o caminho antes de continuar.'); pause(); return
    nome_original = nome_arquivo_seguro(caminho_path.name)
    ja_importado = fetchone('SELECT COUNT(*) AS c FROM oc_itens WHERE observacao LIKE ?', (f'%Arquivo: %{nome_original}%',))
    if ja_importado and int(ja_importado['c'] or 0) > 0:
        warn(f'Esta OC aparenta já ter sido importada ({ja_importado["c"]} itens encontrados).')
        if not confirm('Importar novamente mesmo assim? Isso pode duplicar quantidades na conferência.', default=False):
            warn('Importação cancelada para evitar duplicidade.')
            pause(); return
    try:
        itens = extrair_oc_pdf_tekinisa(caminho_path)
    except Exception as e:
        warn(str(e)); pause(); return
    if not itens:
        warn('Nenhum item encontrado. Confira se o PDF é a Solicitação de Compra padrão Tekinisa.')
        pause(); return
    docs = sorted(set(i['documento'] for i in itens))
    entregas = sorted(set(i['data_entrega'] for i in itens if i['data_entrega']))
    usos = sorted(set(i['data_utilizacao'] for i in itens if i['data_utilizacao']))
    cats = [(docs[0] if docs else 'OC', len(itens)), ('Solicitações/blocos', len(docs)), ('Datas de entrega', len(entregas)), ('Datas de utilização', len(usos))]
    tabela('PRÉVIA DA OC', ['Informação', 'Total'], cats)
    print(f"\nArquivo: {Path(caminho).name}")
    print(f"Entrega: {iso_to_br(entregas[0])} a {iso_to_br(entregas[-1])}" if entregas else 'Entrega não identificada')
    print(f"Utilização: {iso_to_br(usos[0])} a {iso_to_br(usos[-1])}" if usos else 'Utilização não identificada')
    print(f"Itens encontrados: {len(itens)}")
    if not confirm('Confirmar importação desta Solicitação de Compra?', default=True):
        warn('Importação cancelada.'); pause(); return
    dest = IMPORT_DIR / f'oc_tekinisa_{agora_stamp()}_{nome_original}'
    try:
        shutil.copy2(caminho, dest)
        arquivo = dest.name
    except Exception:
        arquivo = nome_original
    sql_import = '''INSERT INTO oc_itens(documento,data_entrega,data_utilizacao,codigo,item,quantidade,unidade,observacao,
                    emissao,versao,filial_codigo,filial_nome,periodo_entrega_inicio,periodo_entrega_fim,solicitacao,sequencia,data_solicitacao,descricao)
                   VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)'''
    dados_import = []
    for x in itens:
        obs = f"Arquivo: {arquivo}; Solicitação: {x['solicitacao']}; Emissão: {x['emissao']}; Filial: {x['filial_codigo']} {x['filial_nome']}"
        dados_import.append((x['documento'], x['data_entrega'], x['data_utilizacao'], x['codigo'], x['item'], x['quantidade'], x['unidade'], obs,
                 x['emissao'], x['versao'], x['filial_codigo'], x['filial_nome'], x['periodo_entrega_inicio'], x['periodo_entrega_fim'],
                 x['solicitacao'], x['sequencia'], x['data_solicitacao'], x['descricao']))
    total = executemany(sql_import, dados_import)
    evento('OC', f'Importada Solicitação de Compra Tekinisa {arquivo} com {total} itens')
    ok(f'Importação concluída: {total} itens da OC salvos.')
    pause()


def listar_oc():
    header('ITENS DE ORDEM DE COMPRA')
    rows = fetchall('SELECT id,documento,codigo,item,quantidade,unidade,data_entrega,data_utilizacao,observacao FROM oc_itens ORDER BY id DESC LIMIT 250')
    dados = [(r['id'], r['documento'], r['codigo'], r['item'], fmt_qtd(r['quantidade'], r['unidade']), r['unidade'], iso_to_br(r['data_entrega']), iso_to_br(r['data_utilizacao'])) for r in rows]
    tabela('OC / SOLICITAÇÕES', ['ID','Documento','Código','Item','Qtd','Un','Entrega','Utilização'], dados, 250)
    pause()


def _oc_rows_periodo(ini=None, fim=None, filtrar_utilizacao=True):
    if ini and fim and filtrar_utilizacao:
        return fetchall('''SELECT * FROM oc_itens WHERE (data_utilizacao BETWEEN ? AND ?) OR data_utilizacao='' OR data_utilizacao IS NULL ORDER BY item''', (ini, fim))
    return fetchall('SELECT * FROM oc_itens ORDER BY item')


def _oc_total_por_item(ini=None, fim=None, filtrar_utilizacao=True):
    rows = _oc_rows_periodo(ini, fim, filtrar_utilizacao)
    agrup = {}
    for r in rows:
        chave = (r['codigo'] or '', _chave_nome(r['item']), r['unidade'] or '')
        if chave not in agrup:
            agrup[chave] = {'codigo': r['codigo'] or '', 'item': r['item'], 'unidade': r['unidade'], 'qtd': 0.0, 'entregas': [], 'utilizacoes': [], 'documentos': [], 'ids': []}
        agrup[chave]['qtd'] += float(r['quantidade'] or 0)
        if r['data_entrega']:
            agrup[chave]['entregas'].append(r['data_entrega'])
        if r['data_utilizacao']:
            agrup[chave]['utilizacoes'].append(r['data_utilizacao'])
        if r['documento'] and r['documento'] not in agrup[chave]['documentos']:
            agrup[chave]['documentos'].append(r['documento'])
        agrup[chave]['ids'].append(r['id'])
    return list(agrup.values())


def _similar(a, b):
    return SequenceMatcher(None, _chave_nome(a), _chave_nome(b)).ratio()


def _achar_oc(item, codigo='', unidade='', oc_totais=None):
    oc_totais = oc_totais if oc_totais is not None else _oc_total_por_item()
    codigo = str(codigo or '').strip()
    item_key = _chave_nome(item)
    unidade = str(unidade or '').upper().strip()
    if codigo:
        for r in oc_totais:
            if str(r.get('codigo') or '').strip() == codigo and (not unidade or not r.get('unidade') or r.get('unidade') == unidade):
                return r, 'CÓDIGO', 1.0
    for r in oc_totais:
        if _chave_nome(r['item']) == item_key and (not unidade or not r.get('unidade') or r.get('unidade') == unidade):
            return r, 'NOME EXATO', 1.0
    candidatos = []
    for r in oc_totais:
        ok_ = _chave_nome(r['item'])
        if item_key and (item_key in ok_ or ok_ in item_key):
            score = 0.90
            if unidade and r.get('unidade') and r.get('unidade') != unidade:
                score -= .15
            candidatos.append((score, r, 'NOME PARECIDO'))
        else:
            score = _similar(item_key, ok_)
            if score >= .72:
                if unidade and r.get('unidade') and r.get('unidade') != unidade:
                    score -= .15
                candidatos.append((score, r, 'SIMILARIDADE'))
    if candidatos:
        candidatos.sort(key=lambda x: x[0], reverse=True)
        score, r, metodo = candidatos[0]
        return r, metodo, score
    return None, '', 0.0


def _classificar_conferencia(necessario, qtd_oc, unidade_analitico, unidade_oc, datas_uso, entregas):
    unidade_ok = (not unidade_oc) or (unidade_analitico == unidade_oc)
    saldo = qtd_oc - necessario
    primeiro_uso = min(datas_uso) if datas_uso else ''
    primeira_entrega = min(entregas) if entregas else ''
    entrega_risco = bool(primeiro_uso and primeira_entrega and primeira_entrega > primeiro_uso)
    if not unidade_ok:
        return 'VERIFICAR UNIDADE', 'MÉDIA', saldo, 'Unidade do analítico diferente da unidade da OC.'
    if entrega_risco:
        return 'ENTREGA EM RISCO', 'ALTA', saldo, 'Data de entrega da OC é posterior ao primeiro consumo previsto.'
    if qtd_oc <= 0:
        return 'FALTANTE NA OC', 'ALTA', -necessario, 'Item necessário no período não encontrado na OC.'
    tolerancia = 0.000001
    if saldo < -tolerancia:
        crit = 'ALTA' if abs(saldo) >= max(necessario * 0.25, 1) else 'MÉDIA'
        return 'QUANTIDADE INSUFICIENTE', crit, saldo, 'Quantidade da OC não cobre a necessidade do período.'
    if saldo > tolerancia:
        return 'OK COM SOBRA', 'BAIXA', saldo, 'OC cobre a necessidade e possui sobra calculada.'
    return 'OK', 'BAIXA', 0.0, 'OC cobre exatamente a necessidade do período.'


def _conferir_compra_necessidade(ini, fim, turnos, filtrar_oc_por_utilizacao=True):
    base = _resultado_necessario_estoque(ini, fim, turnos)
    oc_totais = _oc_total_por_item(ini, fim, filtrar_oc_por_utilizacao)
    usados_oc = set()
    resultado = []
    for x in base:
        oc, metodo, score = _achar_oc(x['item'], x.get('codigo_analitico') or '', x['unidade'], oc_totais)
        qtd_oc = float(oc['qtd'] or 0) if oc else 0.0
        item_oc = oc['item'] if oc else ''
        unidade_oc = oc['unidade'] if oc else ''
        entregas = sorted(set(oc['entregas'])) if oc else []
        utilizacoes = sorted(set(oc['utilizacoes'])) if oc else []
        if oc:
            for oid in oc.get('ids', []):
                usados_oc.add(oid)
        status, criticidade, saldo_oc, obs = _classificar_conferencia(x['necessario'], qtd_oc, x['unidade'], unidade_oc, x['datas_uso'], entregas)
        resultado.append({**x, 'item_oc': item_oc, 'codigo_oc': oc['codigo'] if oc else '', 'unidade_oc': unidade_oc, 'qtd_oc': qtd_oc, 'saldo_oc': saldo_oc,
                          'status_compra': status, 'criticidade_compra': criticidade, 'metodo': metodo, 'score': score,
                          'entregas': entregas, 'utilizacoes_oc': utilizacoes, 'observacao_compra': obs})
    oc_sem_uso = []
    for r in oc_totais:
        if not any(oid in usados_oc for oid in r.get('ids', [])):
            oc_sem_uso.append({
                'status_compra': 'ITEM NA OC SEM USO NO PERÍODO', 'criticidade_compra': 'VERIFICAR', 'categoria': 'OC',
                'item': '', 'codigo_analitico': '', 'unidade': r['unidade'], 'necessario': 0.0, 'estoque': 0.0, 'saldo': 0.0,
                'produto': '', 'codigo': '', 'primeira_data': '', 'ultima_data': '', 'datas_uso': [],
                'item_oc': r['item'], 'codigo_oc': r['codigo'], 'unidade_oc': r['unidade'], 'qtd_oc': r['qtd'], 'saldo_oc': r['qtd'],
                'metodo': 'OC SEM CORRESPONDÊNCIA', 'score': 0.0, 'entregas': sorted(set(r['entregas'])), 'utilizacoes_oc': sorted(set(r['utilizacoes'])),
                'observacao_compra': 'Item está na OC, mas não apareceu na necessidade analítica do período/turno selecionado.'
            })
    return resultado, oc_sem_uso


def _gerar_excel_conferencia_compra(ini, fim, turnos, resultado, oc_sem_uso):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    arq = EXCEL_DIR / f'conferencia_compra_necessidade_{agora_stamp()}.xlsx'
    wb = Workbook()
    azul = '1F4E78'; vermelho = '8B0000'; laranja = 'F4B183'; verde = '70AD47'

    ws = wb.active
    ws.title = 'Resumo Executivo'
    total = len(resultado)
    faltantes = len([r for r in resultado if r['status_compra'] in ('FALTANTE NA OC','QUANTIDADE INSUFICIENTE','ENTREGA EM RISCO','VERIFICAR UNIDADE')])
    ok_count = len([r for r in resultado if r['status_compra'] in ('OK','OK NO LIMITE','OK COM SOBRA')])
    ws.append([APP_NAME])
    ws.append([f'Conferência Compra x Necessidade - {iso_to_br(ini)} a {iso_to_br(fim)} - {assinatura_relatorio()}'])
    ws.append([f'Turnos: {_turnos_txt(turnos)}'])
    ws.append([])
    for linha in [['Itens necessários', total], ['Itens com atenção', faltantes], ['Itens OK / OK com sobra', ok_count], ['Itens na OC sem uso', len(oc_sem_uso)], ['Emissão', assinatura_relatorio()]]:
        ws.append(linha)
    ws.column_dimensions['A'].width = 28; ws.column_dimensions['B'].width = 20

    def sheet(nome, rows):
        wsx = wb.create_sheet(nome[:31])
        headers = ['Status','Criticidade','Categoria','Código Analítico','Item Analítico','Necessário','Un','Datas de consumo','Código OC','Item OC','Qtd OC','Un OC','Entrega OC','Utilização OC','Falta/Sobra','Método','Observação']
        wsx.append(headers)
        for c in wsx[1]:
            c.font=Font(bold=True,color='FFFFFF'); c.fill=PatternFill('solid', fgColor=azul); c.alignment=Alignment(horizontal='center')
        for r in rows:
            wsx.append([r['status_compra'], r['criticidade_compra'], r.get('categoria',''), r.get('codigo_analitico',''), r.get('item',''),
                        float(r.get('necessario') or 0), r.get('unidade',''), ', '.join(iso_to_br(d) for d in r.get('datas_uso',[])),
                        r.get('codigo_oc',''), r.get('item_oc',''), float(r.get('qtd_oc') or 0), r.get('unidade_oc',''),
                        ', '.join(iso_to_br(d) for d in r.get('entregas',[])), ', '.join(iso_to_br(d) for d in r.get('utilizacoes_oc',[])),
                        float(r.get('saldo_oc') or 0), r.get('metodo',''), r.get('observacao_compra','')])
        widths=[24,14,18,18,52,14,8,45,18,52,14,8,28,28,14,18,55]
        for i,w in enumerate(widths,1):
            wsx.column_dimensions[chr(64+i)].width=w
        wsx.freeze_panes='A2'
        return wsx

    sheet('Todos os Itens', resultado)
    sheet('Faltantes e Divergentes', [r for r in resultado if r['status_compra'] in ('FALTANTE NA OC','QUANTIDADE INSUFICIENTE','ENTREGA EM RISCO','VERIFICAR UNIDADE')])
    sheet('Itens OK', [r for r in resultado if r['status_compra'] in ('OK','OK NO LIMITE')])
    sheet('OK com Sobra', [r for r in resultado if r['status_compra'] == 'OK COM SOBRA'])
    sheet('OC sem Uso', oc_sem_uso)
    wb.save(arq)
    return arq


def _gerar_pdf_conferencia_compra(ini, fim, turnos, resultado, oc_sem_uso):
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.units import cm
    arq = PDF_DIR / f'conferencia_compra_necessidade_{agora_stamp()}.pdf'
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='TituloCentro', parent=styles['Title'], alignment=TA_CENTER, fontSize=14, leading=17))
    styles.add(ParagraphStyle(name='Secao', parent=styles['Heading2'], fontSize=11, leading=13, spaceBefore=6, spaceAfter=5))
    styles.add(ParagraphStyle(name='Peq', parent=styles['BodyText'], fontSize=6.8, leading=8))
    def p(txt, style='Peq'):
        return Paragraph(escape(str(txt)), styles[style])
    doc = SimpleDocTemplate(str(arq), pagesize=landscape(A4), leftMargin=.6*cm, rightMargin=.6*cm, topMargin=.7*cm, bottomMargin=1*cm)
    story=[]
    story.append(p('RELATÓRIO DE CONFERÊNCIA DE COMPRA', 'TituloCentro'))
    story.append(p(f'Analítico x Solicitação de Compra Tekinisa | {iso_to_br(ini)} a {iso_to_br(fim)} | {_turnos_txt(turnos)}', 'Peq'))
    story.append(Spacer(1, 7))
    falt = [r for r in resultado if r['status_compra'] in ('FALTANTE NA OC','QUANTIDADE INSUFICIENTE','ENTREGA EM RISCO','VERIFICAR UNIDADE')]
    resumo = [['Indicador','Quantidade'], ['Itens necessários', str(len(resultado))], ['Itens com atenção', str(len(falt))], ['Itens na OC sem uso', str(len(oc_sem_uso))], ['Emissão', assinatura_relatorio()]]
    tb = Table([[p(a), p(b)] for a,b in resumo], colWidths=[7*cm, 5*cm])
    tb.setStyle(TableStyle([('GRID',(0,0),(-1,-1),.25,colors.grey),('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1F4E78')),('TEXTCOLOR',(0,0),(-1,0),colors.white)]))
    story.append(tb); story.append(Spacer(1,8))

    def tabela_secao(titulo, rows):
        story.append(p(titulo, 'Secao'))
        data = [[p('Status'), p('Crit.'), p('Item Analítico'), p('Nec.'), p('Un'), p('Datas consumo'), p('Item OC'), p('Qtd OC'), p('Entrega'), p('Falta/Sobra')]]
        for r in rows:
            data.append([p(r['status_compra']), p(r['criticidade_compra']), p(r.get('item','')[:45]), p(fmt_qtd(r.get('necessario',0), r.get('unidade',''))), p(r.get('unidade','')), p(', '.join(iso_to_br(d) for d in r.get('datas_uso',[]))[:35]), p(r.get('item_oc','')[:42]), p(fmt_qtd(r.get('qtd_oc',0), r.get('unidade_oc',''))), p(', '.join(iso_to_br(d) for d in r.get('entregas',[]))[:26]), p(fmt_qtd(r.get('saldo_oc',0), r.get('unidade','')))])
        if len(data)==1:
            data.append([p('OK'),p('-'),p('Nenhum item nesta seção'),p('-'),p('-'),p('-'),p('-'),p('-'),p('-'),p('-')])
        t=Table(data, repeatRows=1, colWidths=[3.1*cm,1.6*cm,7.0*cm,2.0*cm,1.0*cm,4.0*cm,6.8*cm,2.0*cm,2.5*cm,2.0*cm])
        t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1F4E78')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('GRID',(0,0),(-1,-1),.25,colors.grey),('FONTSIZE',(0,0),(-1,-1),6.2),('VALIGN',(0,0),(-1,-1),'TOP')]))
        story.append(t); story.append(Spacer(1,6))
    tabela_secao('1. ITENS FALTANTES / DIVERGENTES / EM RISCO', falt)
    story.append(PageBreak())
    tabela_secao('2. ITENS OK / COBERTOS PELA OC', [r for r in resultado if r['status_compra'] in ('OK','OK NO LIMITE','OK COM SOBRA')])
    story.append(PageBreak())
    tabela_secao('3. ITENS NA OC SEM USO NO PERÍODO SELECIONADO', oc_sem_uso)
    def footer(canvas, doc):
        canvas.saveState(); canvas.setFont('Helvetica',7); canvas.drawRightString(28.6*cm, .45*cm, f'{assinatura_relatorio()} | Página {doc.page}'); canvas.restoreState()
    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    return arq


def conferir_compra_necessidade():
    header('CONFERIR COMPRA X NECESSIDADE')
    info('Use esta função após importar o Analítico e importar a Solicitação de Compra Tekinisa.')
    ini, fim = pedir_periodo('Data inicial do uso', 'Data final do uso')
    turnos = escolher_turnos()
    filtrar = confirm('Considerar apenas itens da OC com data de utilização dentro do período?', default=True)
    resultado, oc_sem_uso = _conferir_compra_necessidade(ini, fim, turnos, filtrar)
    if not resultado:
        warn('Nenhum item analítico encontrado para o período. Importe o analítico primeiro.')
        pause(); return
    dados = [(r['status_compra'], r['criticidade_compra'], r['item'], fmt_qtd(r['necessario'], r['unidade']), r['unidade'], r['item_oc'], fmt_qtd(r['qtd_oc'], r['unidade_oc']), fmt_qtd(r['saldo_oc'], r['unidade'])) for r in resultado if r['status_compra'] not in ('OK','OK NO LIMITE','OK COM SOBRA')]
    tabela('ITENS FALTANTES / DIVERGENTES', ['Status','Crit.','Item necessário','Necess.','Un','Item OC','Qtd OC','Falta/Sobra'], dados, 200)
    if confirm('Gerar relatório completo PDF + Excel?', default=True):
        try:
            excel = _gerar_excel_conferencia_compra(ini, fim, turnos, resultado, oc_sem_uso)
            pdf = _gerar_pdf_conferencia_compra(ini, fim, turnos, resultado, oc_sem_uso)
            ok(f'Excel gerado: {excel}')
            ok(f'PDF gerado: {pdf}')
        except Exception as e:
            warn(f'Erro ao gerar relatório: {e}')
    evento('OC', f'Conferência Compra x Necessidade executada para {iso_to_br(ini)} a {iso_to_br(fim)}')
    pause()


def conferir_analitico_estoque_oc():
    header('CONFERÊNCIA ANALÍTICO × ESTOQUE × OC')
    ini, fim = pedir_periodo('Data inicial', 'Data final')
    turnos = escolher_turnos()
    base = _resultado_necessario_estoque(ini, fim, turnos)
    if not base:
        warn('Nenhum dado analítico encontrado. Importe a requisição analítica primeiro.')
        pause(); return
    oc_totais = _oc_total_por_item(ini, fim, True)
    resultado=[]
    for x in base:
        oc, metodo, score = _achar_oc(x['item'], x.get('codigo_analitico') or '', x['unidade'], oc_totais)
        qtd_oc = float(oc['qtd'] or 0) if oc else 0.0
        previsto = float(x['estoque']) + qtd_oc
        saldo_final = previsto - float(x['necessario'])
        if saldo_final < 0:
            status = 'FALTANTE APÓS OC'; criticidade = 'ALTA'
        elif not oc and x['saldo'] < 0:
            status = 'SEM OC CADASTRADA'; criticidade = 'ALTA'
        elif saldo_final <= max(float(x['necessario']) * .10, 1):
            status = 'COBRE NO LIMITE'; criticidade = 'MÉDIA'
        else:
            status = 'COBRE COM OC/ESTOQUE'; criticidade = 'BAIXA'
        resultado.append({**x, 'oc': qtd_oc, 'previsto': previsto, 'saldo_final': saldo_final, 'status_oc': status, 'criticidade_oc': criticidade, 'item_oc': oc['item'] if oc else ''})
    dados=[(r['status_oc'],r['criticidade_oc'],r['categoria'],r['item'],fmt_qtd(r['necessario'],r['unidade']),fmt_qtd(r['estoque'],r['unidade']),fmt_qtd(r['oc'],r['unidade']),fmt_qtd(r['saldo_final'],r['unidade']),r['unidade']) for r in resultado]
    tabela('RESULTADO GERAL', ['Status','Crit.','Categoria','Item','Necess.','Estoque','OC','Saldo Final','Un'], dados, 300)
    if confirm('Gerar Excel da conferência com OC?', default=True):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            arq = EXCEL_DIR / f'conferencia_analitico_estoque_oc_{agora_stamp()}.xlsx'
            wb=Workbook(); ws=wb.active; ws.title='Conferência OC'
            ws.append([APP_NAME]); ws.append([f'Analítico x Estoque x OC - {iso_to_br(ini)} a {iso_to_br(fim)} - {assinatura_relatorio()}']); ws.append([])
            headers=['Status','Criticidade','Categoria','Item Analítico','Item OC','Necessário','Estoque','OC','Previsto','Saldo Final','Unidade','Datas consumo']
            ws.append(headers)
            for c in ws[4]: c.font=Font(bold=True,color='FFFFFF'); c.fill=PatternFill('solid', fgColor='1F4E78'); c.alignment=Alignment(horizontal='center')
            for r in resultado: ws.append([r['status_oc'],r['criticidade_oc'],r['categoria'],r['item'],r['item_oc'],r['necessario'],r['estoque'],r['oc'],r['previsto'],r['saldo_final'],r['unidade'], ', '.join(iso_to_br(d) for d in r['datas_uso'])])
            for i,w in enumerate([22,14,20,50,40,14,14,14,14,14,10,45],1): ws.column_dimensions[chr(64+i)].width=w
            wb.save(arq); ok(f'Excel gerado: {arq}')
        except Exception as e: warn(f'Erro ao gerar Excel: {e}')
    evento('OC', f'Conferência Analítico x Estoque x OC executada para {iso_to_br(ini)} a {iso_to_br(fim)}')
    pause()


def menu_oc():
    while True:
        op = menu('COMPRA E CONFERÊNCIA', [
            ('1','Importar Solicitação de Compra Tekinisa PDF'),
            ('2','Conferir Necessários x Ordem de Compra'),
            ('3','Relatório de Itens Faltantes e Divergentes'),
            ('4','Consulta auxiliar: Necessário x Estoque x OC'),
            ('5','Cadastrar item manualmente'),
            ('6','Importar OC em CSV'),
            ('7','Listar itens cadastrados'),
            ('0','Voltar')
        ])
        if op=='1': importar_oc_pdf()
        elif op=='2': conferir_compra_necessidade()
        elif op=='3': conferir_compra_necessidade()
        elif op=='4': conferir_analitico_estoque_oc()
        elif op=='5': cadastrar_item_oc()
        elif op=='6': importar_oc_csv()
        elif op=='7': listar_oc()
        elif op=='0': break
        else: warn('Opção inválida.'); pause()
