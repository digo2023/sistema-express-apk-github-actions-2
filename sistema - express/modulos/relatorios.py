from configuracao import EXCEL_DIR, PDF_DIR, APP_NAME, agora_stamp, assinatura_relatorio
from modulos.banco_dados import fetchall
from modulos.interface import header, ok, warn, tabela, pause, menu, executar_seguro
from modulos.utilitarios import fmt_qtd, iso_to_br


def dados_estoque(apenas_com_saldo=False):
    rows = fetchall('''SELECT p.codigo,p.nome,p.categoria,p.unidade,p.estoque_minimo,p.codigo_barras,
                              e.localizacao,e.lote,e.validade,e.quantidade,
                              COALESCE((SELECT SUM(x.quantidade) FROM estoque_locais x WHERE x.produto_id=p.id),0) AS total
                       FROM produtos p LEFT JOIN estoque_locais e ON e.produto_id=p.id
                       WHERE p.ativo=1
                       ORDER BY p.nome COLLATE NOCASE, e.localizacao COLLATE NOCASE, e.validade, e.lote''')
    if apenas_com_saldo:
        rows = [r for r in rows if float(r['total'] or 0) > 0]
    return rows


def estoque_tela(apenas_com_saldo=False):
    header('ESTOQUE COM SALDO' if apenas_com_saldo else 'ESTOQUE COMPLETO')
    rows = dados_estoque(apenas_com_saldo)
    dados=[(r['codigo'],r['nome'],r['categoria'],r['codigo_barras'] or '',r['localizacao'] or '-',r['lote'],iso_to_br(r['validade']),fmt_qtd(r['quantidade'] or 0,r['unidade']),r['unidade'],fmt_qtd(r['total'],r['unidade']),fmt_qtd(r['estoque_minimo'],r['unidade'])) for r in rows]
    tabela('ESTOQUE POR LOCALIZAÇÃO', ['Código','Produto','Categoria','Cód. Barras','Local','Lote','Validade','Qtd Local','Un','Total','Mínimo'], dados, 1000000)
    pause()


def excel_estoque(apenas_com_saldo=False):
    rows=dados_estoque(apenas_com_saldo)
    nome = 'estoque_com_saldo' if apenas_com_saldo else 'estoque_completo'
    arquivo=EXCEL_DIR / f'{nome}_{agora_stamp()}.xlsx'
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb=Workbook(); ws=wb.active; ws.title='Estoque'
        ws.append([APP_NAME]); ws.append([('Estoque com saldo' if apenas_com_saldo else 'Estoque completo') + f' - {assinatura_relatorio()}']); ws.append([])
        headers=['Código','Produto','Categoria','Código de barras','Localização','Lote','Validade','Qtd Local','Un','Total Produto','Mínimo']
        ws.append(headers)
        for c in ws[4]:
            c.font=Font(bold=True, color='FFFFFF'); c.fill=PatternFill('solid', fgColor='1F4E78'); c.alignment=Alignment(horizontal='center')
        for r in rows:
            ws.append([r['codigo'],r['nome'],r['categoria'],r['codigo_barras'] or '',r['localizacao'] or '',r['lote'] or '',iso_to_br(r['validade']),r['quantidade'] or 0,r['unidade'],r['total'],r['estoque_minimo']])
        widths=[16,38,18,22,22,16,14,14,10,14,14]
        for i,w in enumerate(widths,1): ws.column_dimensions[chr(64+i)].width=w
        wb.save(arquivo); ok(f'Excel gerado: {arquivo}')
    except Exception as e: warn(f'Erro ao gerar Excel: {e}')
    pause()


def excel_movimentacoes():
    rows=fetchall('''SELECT m.*,p.codigo,p.nome,p.categoria FROM movimentacoes m JOIN produtos p ON p.id=m.produto_id ORDER BY m.id DESC''')
    arquivo=EXCEL_DIR / f'movimentacoes_express_{agora_stamp()}.xlsx'
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb=Workbook(); ws=wb.active; ws.title='Movimentações'
        ws.append([APP_NAME]); ws.append([f'Movimentações - {assinatura_relatorio()}']); ws.append([])
        headers=['ID','Código','Produto','Categoria','Tipo','Qtd','Un','Local','Lote','Validade','Data Mov','Data Consumo','Turno','Motivo','Obs','Antes','Depois']
        ws.append(headers)
        for c in ws[4]: c.font=Font(bold=True,color='FFFFFF'); c.fill=PatternFill('solid', fgColor='1F4E78'); c.alignment=Alignment(horizontal='center')
        for r in rows:
            ws.append([r['id'],r['codigo'],r['nome'],r['categoria'],r['tipo'],r['quantidade'],r['unidade'],r['localizacao'],r['lote'],iso_to_br(r['validade']),iso_to_br(r['data_movimento']),iso_to_br(r['data_consumo']),r['turno'],r['motivo'],r['observacao'],r['estoque_antes'],r['estoque_depois']])
        for i,w in enumerate([8,16,36,18,18,12,8,22,14,14,14,14,10,20,40,12,12],1): ws.column_dimensions[chr(64+i) if i<=26 else 'A'].width=w
        wb.save(arquivo); ok(f'Excel gerado: {arquivo}')
    except Exception as e: warn(f'Erro ao gerar Excel: {e}')
    pause()


def pdf_estoque(apenas_com_saldo=False):
    rows=dados_estoque(apenas_com_saldo)
    nome = 'estoque_com_saldo' if apenas_com_saldo else 'estoque_completo'
    arquivo=PDF_DIR / f'{nome}_{agora_stamp()}.pdf'
    try:
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
        styles=getSampleStyleSheet(); doc=SimpleDocTemplate(str(arquivo), pagesize=landscape(A4), leftMargin=.7*cm,rightMargin=.7*cm,topMargin=.7*cm,bottomMargin=1*cm)
        titulo = 'Estoque com saldo' if apenas_com_saldo else 'Estoque completo'
        story=[Paragraph(APP_NAME, styles['Title']), Paragraph(f'{titulo} - {assinatura_relatorio()}', styles['Normal']), Spacer(1,8)]
        data=[['Código','Produto','Categoria','Local','Val','Qtd','Un','Total','Mín']]
        for r in rows:
            data.append([r['codigo'],r['nome'][:34],r['categoria'],(r['localizacao'] or '-')[:20],iso_to_br(r['validade']),fmt_qtd(r['quantidade'] or 0,r['unidade']),r['unidade'],fmt_qtd(r['total'],r['unidade']),fmt_qtd(r['estoque_minimo'],r['unidade'])])
        table=Table(data, repeatRows=1, colWidths=[2.3*cm,6.3*cm,3*cm,4*cm,2*cm,2*cm,1.2*cm,2*cm,2*cm])
        table.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1F4E78')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('GRID',(0,0),(-1,-1),0.25,colors.grey),('FONTSIZE',(0,0),(-1,-1),7.5),('VALIGN',(0,0),(-1,-1),'MIDDLE')]))
        story.append(table)
        def footer(canvas, doc):
            canvas.saveState(); canvas.setFont('Helvetica',8); canvas.drawRightString(28.5*cm, .45*cm, assinatura_relatorio()); canvas.restoreState()
        doc.build(story, onFirstPage=footer, onLaterPages=footer); ok(f'PDF gerado: {arquivo}')
    except Exception as e: warn(f'Erro ao gerar PDF. Instale reportlab. Detalhe: {e}')
    pause()


def relatorio_diario_movimentacao():
    from modulos.utilitarios import pedir_data
    data = pedir_data('Data do relatório')
    rows=fetchall('''SELECT m.*,p.codigo,p.nome,p.categoria FROM movimentacoes m JOIN produtos p ON p.id=m.produto_id
                     WHERE m.data_movimento=? OR m.data_consumo=? ORDER BY m.tipo,p.nome''', (data,data))
    header('RELATÓRIO DIÁRIO')
    dados=[(r['tipo'],r['codigo'],r['nome'],fmt_qtd(r['quantidade'],r['unidade']),r['unidade'],r['localizacao'],r['turno'],r['motivo'],r['observacao']) for r in rows]
    tabela(f'MOVIMENTAÇÕES DO DIA {iso_to_br(data)}', ['Tipo','Código','Produto','Qtd','Un','Local','Turno','Motivo','Obs'], dados, 1000000)
    if not rows:
        warn('Nenhuma movimentação encontrada para esta data.')
    pause()


def menu_relatorios():
    while True:
        op=menu('RELATÓRIOS', [
            ('1','Ver estoque completo na tela'),
            ('2','Ver estoque com saldo na tela'),
            ('3','Gerar estoque completo PDF + Excel'),
            ('4','Gerar estoque com saldo PDF + Excel'),
            ('5','Gerar Excel de movimentações'),
            ('6','Relatório diário na tela'),
            ('0','Voltar')])
        if op=='1': executar_seguro(estoque_tela, False)
        elif op=='2': executar_seguro(estoque_tela, True)
        elif op=='3': executar_seguro(pdf_estoque, False); executar_seguro(excel_estoque, False)
        elif op=='4': executar_seguro(pdf_estoque, True); executar_seguro(excel_estoque, True)
        elif op=='5': executar_seguro(excel_movimentacoes)
        elif op=='6': executar_seguro(relatorio_diario_movimentacao)
        elif op=='0': break
        else: warn('Opção inválida.'); pause()
