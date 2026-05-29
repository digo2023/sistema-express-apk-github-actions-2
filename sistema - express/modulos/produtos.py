from pathlib import Path
from collections import OrderedDict
from configuracao import CATEGORIAS, UNIDADES, PDF_DIR, EXCEL_DIR, agora_stamp, assinatura_relatorio
from modulos.banco_dados import execute, executemany, fetchall, fetchone, evento
from modulos.interface import header, ask, ask_float, confirm, ok, warn, tabela, pause, menu, info, executar_seguro
from modulos.utilitarios import fmt_qtd, estoque_total, locais_produto, iso_to_br


def escolher_lista(nome, opcoes):
    print(f"\n{nome}:")
    for i, item in enumerate(opcoes, 1): print(f"[{i}] {item}")
    while True:
        op = ask(f"Escolha {nome.lower()}", required=True)
        if op.isdigit() and 1 <= int(op) <= len(opcoes): return opcoes[int(op)-1]
        valor = op.upper().strip()
        if valor in opcoes: return valor
        warn('Opção inválida.')


def cadastrar_produto_fluxo(nome_sugerido=''):
    header('CADASTRO DE PRODUTO')
    print('CADASTRO INTELIGENTE DE PRODUTO\n')
    codigo = ask('Código do produto', upper=True, required=True)
    ja = fetchone('SELECT * FROM produtos WHERE codigo=?', (codigo,))
    if ja:
        warn('Código já cadastrado. Retornando produto existente.'); return ja
    nome = ask('Nome do produto', default=nome_sugerido.upper().replace('_',' ') if nome_sugerido else None, upper=True, required=True)
    categoria = escolher_lista('Categoria', CATEGORIAS)
    unidade = escolher_lista('Unidade', UNIDADES)
    minimo = ask_float('Estoque mínimo', default=0, min_value=0)
    codigo_barras = ask('Código de barras / EAN (opcional)', default='')
    if codigo_barras:
        usado = fetchone('SELECT codigo,nome FROM produtos WHERE codigo_barras=? AND ativo=1', (codigo_barras,))
        if usado:
            warn(f'Código de barras já vinculado ao produto: {usado["codigo"]} - {usado["nome"]}')
            pause(); return None
    obs = ask('Observação', default='')
    print('\nCONFIRA ANTES DE SALVAR:')
    print(f'Código: {codigo}\nProduto: {nome}\nCategoria: {categoria}\nUnidade: {unidade}\nMínimo: {fmt_qtd(minimo, unidade)}\nCódigo de barras: {codigo_barras or '-'}\nObs: {obs}')
    if not confirm('Salvar produto?', default=True):
        warn('Cadastro cancelado.'); pause(); return None
    pid = execute('''INSERT INTO produtos(codigo,nome,categoria,unidade,estoque_minimo,observacao,codigo_barras) VALUES(?,?,?,?,?,?,?)''',
                  (codigo,nome,categoria,unidade,minimo,obs,codigo_barras))
    evento('PRODUTO', f'Produto cadastrado: {codigo} - {nome}')
    ok('Produto cadastrado com sucesso.')
    return fetchone('SELECT * FROM produtos WHERE id=?', (pid,))


def listar_produtos():
    header('PRODUTOS')
    rows = fetchall('''SELECT p.*, COALESCE((SELECT SUM(e.quantidade) FROM estoque_locais e WHERE e.produto_id=p.id),0) AS total
                       FROM produtos p WHERE ativo=1 ORDER BY p.nome''')
    dados = [(r['codigo'], r['nome'], r['categoria'], r['codigo_barras'] or '', fmt_qtd(r['total'], r['unidade']), r['unidade'], fmt_qtd(r['estoque_minimo'], r['unidade'])) for r in rows]
    tabela('CADASTRO DE PRODUTOS', ['Código','Produto','Categoria','Cód. Barras','Estoque','Un','Mínimo'], dados, max_rows=1000000)
    pause()


def pesquisar_produto():
    from modulos.utilitarios import selecionar_produto
    header('PESQUISAR PRODUTO')
    p = selecionar_produto()
    if not p: return
    total = estoque_total(p['id'])
    print(f"\nProduto: {p['nome']}\nCódigo: {p['codigo']}\nCategoria: {p['categoria']}\nUnidade: {p['unidade']}\nCódigo de barras: {p['codigo_barras'] or '-'}\nEstoque total: {fmt_qtd(total,p['unidade'])}\nMínimo: {fmt_qtd(p['estoque_minimo'],p['unidade'])}\nObservação: {p['observacao']}")
    locais = locais_produto(p['id'])
    dados = [(r['localizacao'], r['lote'], iso_to_br(r['validade']), fmt_qtd(r['quantidade'], p['unidade'])) for r in locais]
    tabela('ESTOQUE POR LOCALIZAÇÃO', ['Local','Lote','Validade','Quantidade'], dados)
    pause()


def desativar_produto():
    from modulos.utilitarios import selecionar_produto
    header('DESATIVAR PRODUTO')
    p = selecionar_produto()
    if not p: return
    if estoque_total(p['id']) > 0:
        warn('Produto possui estoque. Recomenda-se zerar por conferência antes de desativar.')
    if confirm(f'Desativar {p["nome"]}?', default=False):
        execute('UPDATE produtos SET ativo=0, atualizado_em=CURRENT_TIMESTAMP WHERE id=?', (p['id'],))
        evento('PRODUTO', f'Produto desativado: {p["codigo"]} - {p["nome"]}')
        ok('Produto desativado.')
    pause()



def _categoria_para_cadastro(categoria_analitico):
    """Converte categorias do analítico para categorias usadas no cadastro de produtos."""
    c = str(categoria_analitico or '').upper().strip()
    mapa = {
        'PROTEINAS': 'PROTEINAS',
        'HORTIFRUTI': 'HORTIFRUTI',
        'DESCARTAVEIS': 'DESCARTAVEL',
        'OPERACIONAIS': 'OUTROS',
        'NAO PERECIVEIS': 'NAO PERECIVEL',
        'NÃO PERECÍVEIS': 'NAO PERECIVEL',
    }
    return mapa.get(c, 'OUTROS')


def _rodape_pdf(canvas, doc):
    canvas.saveState()
    canvas.setFont('Helvetica', 8)
    canvas.drawRightString(570, 15, assinatura_relatorio())
    canvas.restoreState()


def _gerar_excel_cadastro_automatico(produtos_novos, produtos_existentes, produtos_duplicidade_nome):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except Exception:
        return None
    EXCEL_DIR.mkdir(parents=True, exist_ok=True)
    caminho = EXCEL_DIR / f'cadastro_automatico_produtos_{agora_stamp()}.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.title = 'Resumo'
    ws.append(['SISTEMA EXCLUSIVO - EXPRESS RESTAURANTES - UNIDADE COLORADO'])
    ws.append(['CADASTRO AUTOMÁTICO DE PRODUTOS'])
    ws.append([assinatura_relatorio()])
    ws.append([])
    ws.append(['Indicador', 'Quantidade'])
    ws.append(['Produtos novos cadastrados', len(produtos_novos)])
    ws.append(['Produtos já existentes por código', len(produtos_existentes)])
    ws.append(['Possíveis duplicidades por nome/unidade', len(produtos_duplicidade_nome)])
    for cell in ws[1]: cell.font = Font(bold=True, size=13)
    for cell in ws[2]: cell.font = Font(bold=True)
    for cell in ws[5]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1F4E78')
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 18

    def add_sheet(nome, linhas):
        aba = wb.create_sheet(nome)
        headers = ['Código', 'Produto', 'Categoria', 'Unidade', 'Observação']
        aba.append(headers)
        for cell in aba[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='1F4E78')
            cell.alignment = Alignment(horizontal='center')
        for r in linhas:
            aba.append([r.get('codigo',''), r.get('nome',''), r.get('categoria',''), r.get('unidade',''), r.get('observacao','')])
        widths = [20, 48, 20, 12, 55]
        for idx, width in enumerate(widths, 1):
            aba.column_dimensions[chr(64+idx)].width = width

    add_sheet('Novos cadastrados', produtos_novos)
    add_sheet('Já existentes', produtos_existentes)
    add_sheet('Possíveis duplicidades', produtos_duplicidade_nome)
    wb.save(caminho)
    return caminho


def _gerar_pdf_cadastro_automatico(produtos_novos, produtos_existentes, produtos_duplicidade_nome):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
    except Exception:
        return None
    PDF_DIR.mkdir(parents=True, exist_ok=True)
    caminho = PDF_DIR / f'cadastro_automatico_produtos_{agora_stamp()}.pdf'
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(str(caminho), pagesize=A4, rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=30)
    story = []
    story.append(Paragraph('<b>SISTEMA EXCLUSIVO - EXPRESS RESTAURANTES - UNIDADE COLORADO</b>', styles['Title']))
    story.append(Paragraph('<b>RELATÓRIO DE CADASTRO AUTOMÁTICO DE PRODUTOS</b>', styles['Heading2']))
    story.append(Paragraph(assinatura_relatorio(), styles['Normal']))
    story.append(Spacer(1, 10))
    resumo = [
        ['Indicador', 'Quantidade'],
        ['Produtos novos cadastrados', str(len(produtos_novos))],
        ['Produtos já existentes por código', str(len(produtos_existentes))],
        ['Possíveis duplicidades por nome/unidade', str(len(produtos_duplicidade_nome))],
    ]
    t = Table(resumo, colWidths=[330, 130])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F4E78')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('GRID', (0,0), (-1,-1), 0.25, colors.grey),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    story.append(t)
    story.append(Spacer(1, 12))

    def tabela_secao(titulo, linhas):
        story.append(Paragraph(f'<b>{titulo}</b>', styles['Heading3']))
        dados = [['Código', 'Produto', 'Categoria', 'Un', 'Observação']]
        for r in linhas:
            dados.append([r.get('codigo',''), r.get('nome','')[:42], r.get('categoria',''), r.get('unidade',''), r.get('observacao','')[:35]])
        if len(dados) == 1:
            dados.append(['-', 'Nenhum registro', '-', '-', '-'])
        tab = Table(dados, colWidths=[82, 190, 82, 35, 105], repeatRows=1)
        tab.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0B7285')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 7),
            ('GRID', (0,0), (-1,-1), 0.25, colors.lightgrey),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ]))
        story.append(tab)
        story.append(Spacer(1, 10))

    tabela_secao('PRODUTOS NOVOS CADASTRADOS', produtos_novos)
    if produtos_duplicidade_nome:
        story.append(PageBreak())
        tabela_secao('POSSÍVEIS DUPLICIDADES POR NOME E UNIDADE', produtos_duplicidade_nome)
    doc.build(story, onFirstPage=_rodape_pdf, onLaterPages=_rodape_pdf)
    return caminho


def cadastro_automatico_produtos():
    """Lê um PDF Tekinisa e cadastra somente produtos ainda inexistentes."""
    header('CADASTRO AUTOMÁTICO')
    info('Esta função lê o arquivo Tekinisa e cadastra automaticamente apenas produtos ainda inexistentes.')
    info('O estoque inicial será 0. A localização será solicitada quando você fizer a primeira entrada do produto.')
    from modulos.arquivos import selecionar_arquivo
    caminho = selecionar_arquivo('analitico', 'CADASTRO AUTOMÁTICO')
    caminho_pdf = Path(caminho)
    if not caminho_pdf.exists() or not caminho_pdf.is_file():
        warn('Arquivo não encontrado. Confira o caminho e tente novamente.')
        pause(); return
    if caminho_pdf.suffix.lower() != '.pdf':
        warn('Informe um arquivo PDF válido.')
        pause(); return
    try:
        from modulos.analitico import extrair_pdf
        registros = extrair_pdf(caminho_pdf)
    except Exception as exc:
        warn(f'Não foi possível ler o arquivo: {exc}')
        pause(); return
    if not registros:
        warn('Nenhum produto foi encontrado no arquivo informado.')
        pause(); return

    unicos = OrderedDict()
    for r in registros:
        codigo = str(r.get('codigo') or '').strip().upper()
        nome = str(r.get('item') or '').strip().upper()
        unidade = str(r.get('unidade') or '').strip().upper() or 'UN'
        if not codigo or not nome:
            continue
        if codigo not in unicos:
            unicos[codigo] = {
                'codigo': codigo,
                'nome': nome,
                'categoria': _categoria_para_cadastro(r.get('categoria')),
                'unidade': unidade,
                'observacao': 'Cadastrado automaticamente pelo arquivo Tekinisa. Aguardando entrada física/localização.',
            }

    if not unicos:
        warn('Nenhum produto válido foi identificado no arquivo.')
        pause(); return

    codigos = list(unicos.keys())
    existentes_codigo = set()
    for i in range(0, len(codigos), 500):
        lote = codigos[i:i+500]
        ph = ','.join('?' for _ in lote)
        for row in fetchall(f'SELECT codigo FROM produtos WHERE codigo IN ({ph})', lote):
            existentes_codigo.add(str(row['codigo']).upper())

    nomes_unidades_existentes = set((str(r['nome']).upper(), str(r['unidade']).upper()) for r in fetchall('SELECT nome, unidade FROM produtos WHERE ativo=1'))
    produtos_novos = []
    produtos_existentes = []
    produtos_duplicidade_nome = []
    for prod in unicos.values():
        if prod['codigo'] in existentes_codigo:
            item = dict(prod)
            item['observacao'] = 'Ignorado: código já cadastrado no sistema.'
            produtos_existentes.append(item)
        else:
            if (prod['nome'], prod['unidade']) in nomes_unidades_existentes:
                item = dict(prod)
                item['observacao'] = 'Atenção: nome e unidade já existem com outro código. Produto será cadastrado pelo código Tekinisa.'
                produtos_duplicidade_nome.append(item)
            produtos_novos.append(prod)

    tabela('PRÉVIA DO CADASTRO AUTOMÁTICO', ['Situação', 'Quantidade'], [
        ('Produtos únicos identificados no arquivo', len(unicos)),
        ('Novos produtos para cadastrar', len(produtos_novos)),
        ('Já existentes por código', len(produtos_existentes)),
        ('Possíveis duplicidades por nome/unidade', len(produtos_duplicidade_nome)),
    ])
    amostra = [(p['codigo'], p['nome'], p['categoria'], p['unidade']) for p in produtos_novos[:25]]
    tabela('AMOSTRA DOS NOVOS PRODUTOS', ['Código', 'Produto', 'Categoria', 'Un'], amostra, max_rows=25)
    if not produtos_novos:
        warn('Nenhum produto novo para cadastrar. O arquivo já está coberto pelo cadastro atual.')
        if confirm('Gerar relatório desta verificação?', default=True):
            pdf = _gerar_pdf_cadastro_automatico(produtos_novos, produtos_existentes, produtos_duplicidade_nome)
            excel = _gerar_excel_cadastro_automatico(produtos_novos, produtos_existentes, produtos_duplicidade_nome)
            if pdf: ok(f'PDF gerado: {pdf}')
            if excel: ok(f'Excel gerado: {excel}')
        pause(); return

    if not confirm(f'Confirmar cadastro de {len(produtos_novos)} produtos novos?', default=True):
        warn('Cadastro automático cancelado. Nenhum produto foi salvo.')
        pause(); return

    dados_insert = [(p['codigo'], p['nome'], p['categoria'], p['unidade'], 0, p['observacao'], '', 1) for p in produtos_novos]
    try:
        executemany("""INSERT OR IGNORE INTO produtos(codigo,nome,categoria,unidade,estoque_minimo,observacao,codigo_barras,ativo)
                       VALUES(?,?,?,?,?,?,?,?)""", dados_insert)
    except Exception as exc:
        warn(f'Falha ao salvar produtos. Operação interrompida com segurança: {exc}')
        pause(); return
    evento('PRODUTO', f'Cadastro automático concluído: {len(produtos_novos)} produtos novos a partir de {caminho_pdf.name}')

    pdf = _gerar_pdf_cadastro_automatico(produtos_novos, produtos_existentes, produtos_duplicidade_nome)
    excel = _gerar_excel_cadastro_automatico(produtos_novos, produtos_existentes, produtos_duplicidade_nome)
    ok(f'Cadastro automático concluído: {len(produtos_novos)} produtos novos cadastrados.')
    if produtos_existentes:
        info(f'{len(produtos_existentes)} produtos já existiam e foram preservados sem duplicar.')
    if produtos_duplicidade_nome:
        warn(f'{len(produtos_duplicidade_nome)} possíveis duplicidades por nome/unidade foram registradas no relatório para conferência.')
    if pdf: ok(f'PDF gerado: {pdf}')
    else: warn('PDF não gerado. Verifique se reportlab está instalado.')
    if excel: ok(f'Excel gerado: {excel}')
    else: warn('Excel não gerado. Verifique se openpyxl está instalado.')
    pause()

def menu_produtos():
    while True:
        op = menu('PRODUTOS', [('1','Cadastrar produto'),('2','Listar produtos'),('3','Pesquisar produto e localizações'),('4','Desativar produto'),('5','Cadastro automático'),('0','Voltar')])
        if op=='1': executar_seguro(cadastrar_produto_fluxo); pause()
        elif op=='2': executar_seguro(listar_produtos)
        elif op=='3': executar_seguro(pesquisar_produto)
        elif op=='4': executar_seguro(desativar_produto)
        elif op=='5': executar_seguro(cadastro_automatico_produtos)
        elif op=='0': break
        else: warn('Opção inválida.'); pause()
